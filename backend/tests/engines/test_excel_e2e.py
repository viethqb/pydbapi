"""End-to-end tests for Excel report engine.

Covers: seed → execute → verify output xlsx for:
- Basic ROWS / SINGLE write modes
- Write headers on/off
- Multi-mapping same sheet (collision auto-shift + gap_rows)
- Multi-mapping different sheets
- Format config: header format, data format, column_widths
- Auto-fit columns with max width
- Wrap text (global)
- Template-level format merged with mapping-level override
- Chunked pagination (auto-paginate)
- Empty result set
- Inactive mapping skipped
"""

import os
import shutil
import tempfile
import uuid
from unittest.mock import patch

import pytest
from openpyxl import load_workbook
from sqlmodel import Session

from app.core.config import settings as _settings
from app.core.security import encrypt_value
from app.engines.excel.executor import (
    ExcelReportExecutor,
)
from app.models_dbapi import DataSource, ProductTypeEnum
from app.models_report import (
    ExecutionStatusEnum,
    ReportExecution,
    ReportModule,
    ReportSheetMapping,
    ReportTemplate,
    SheetWriteModeEnum,
)

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _create_ds(db: Session, name: str) -> DataSource:
    ds = DataSource(
        name=name,
        product_type=ProductTypeEnum.POSTGRES,
        host=_settings.POSTGRES_SERVER,
        port=_settings.POSTGRES_PORT,
        database=_settings.POSTGRES_DB,
        username=_settings.POSTGRES_USER,
        password=encrypt_value(_settings.POSTGRES_PASSWORD),
        is_active=True,
    )
    db.add(ds)
    db.commit()
    db.refresh(ds)
    return ds


@pytest.fixture()
def sql_ds(db: Session):
    """Real Postgres datasource for running SQL."""
    return _create_ds(db, f"sql-{uuid.uuid4().hex[:8]}")


@pytest.fixture()
def minio_ds(db: Session):
    """Dummy datasource used as minio placeholder (not actually called)."""
    return _create_ds(db, f"minio-{uuid.uuid4().hex[:8]}")


@pytest.fixture()
def work_dir():
    d = tempfile.mkdtemp(prefix="report_test_")
    yield d
    shutil.rmtree(d, ignore_errors=True)


def _make_module(db: Session, sql_ds, minio_ds) -> ReportModule:
    mod = ReportModule(
        name=f"mod-{uuid.uuid4().hex[:8]}",
        minio_datasource_id=minio_ds.id,
        sql_datasource_id=sql_ds.id,
        default_template_bucket="tpl",
        default_output_bucket="out",
    )
    db.add(mod)
    db.commit()
    db.refresh(mod)
    return mod


def _make_template(db: Session, module: ReportModule, **kwargs) -> ReportTemplate:
    tpl = ReportTemplate(
        report_module_id=module.id,
        name=kwargs.pop("name", f"tpl-{uuid.uuid4().hex[:8]}"),
        output_bucket="out",
        template_path="",
        **kwargs,
    )
    db.add(tpl)
    db.commit()
    db.refresh(tpl)
    return tpl


def _make_mapping(db: Session, tpl: ReportTemplate, **kwargs) -> ReportSheetMapping:
    m = ReportSheetMapping(
        report_template_id=tpl.id,
        sort_order=kwargs.pop("sort_order", 0),
        sheet_name=kwargs.pop("sheet_name", "Sheet1"),
        start_cell=kwargs.pop("start_cell", "A1"),
        write_mode=kwargs.pop("write_mode", SheetWriteModeEnum.ROWS),
        write_headers=kwargs.pop("write_headers", False),
        gap_rows=kwargs.pop("gap_rows", 0),
        format_config=kwargs.pop("format_config", None),
        sql_content=kwargs.pop("sql_content", "SELECT 1 AS val"),
        is_active=kwargs.pop("is_active", True),
        **kwargs,
    )
    db.add(m)
    db.commit()
    db.refresh(m)
    return m


def _make_execution(db: Session, tpl: ReportTemplate) -> ReportExecution:
    exc = ReportExecution(
        report_template_id=tpl.id,
        status=ExecutionStatusEnum.PENDING,
    )
    db.add(exc)
    db.commit()
    db.refresh(exc)
    return exc


def _run_executor(db, module, tpl, mappings, execution, params=None):
    """Run executor with mocked MinIO (no real upload/download)."""
    with patch("app.engines.excel.executor.get_minio_client"), \
         patch("app.engines.excel.executor.download_file"), \
         patch("app.engines.excel.executor.upload_file") as mock_upload:

        captured = {}

        def _capture_upload(_client, _bucket, _path, local_path, **_kwargs):
            captured["path"] = local_path
            out = os.path.join(tempfile.gettempdir(), f"test_output_{uuid.uuid4().hex}.xlsx")
            shutil.copy2(local_path, out)
            captured["copy"] = out

        mock_upload.side_effect = _capture_upload

        executor = ExcelReportExecutor()
        result = executor.execute(db, module, tpl, mappings, execution, params)
        return result, captured.get("copy")


def _load_output(path: str):
    return load_workbook(path)


# ---------------------------------------------------------------------------
# 1. Basic ROWS mode
# ---------------------------------------------------------------------------


class TestBasicRows:
    def test_rows_no_headers(self, db, sql_ds, minio_ds):
        mod = _make_module(db, sql_ds, minio_ds)
        tpl = _make_template(db, mod)
        m = _make_mapping(db, tpl, sql_content="SELECT 1 AS id, 'Alice' AS name UNION ALL SELECT 2, 'Bob'")
        exc = _make_execution(db, tpl)

        result, output = _run_executor(db, mod, tpl, [m], exc)
        assert result.status == ExecutionStatusEnum.SUCCESS
        assert output is not None

        wb = _load_output(output)
        ws = wb["Sheet1"]
        assert ws["A1"].value == 1
        assert ws["B1"].value == "Alice"
        assert ws["A2"].value == 2
        assert ws["B2"].value == "Bob"
        wb.close()
        os.unlink(output)

    def test_rows_with_headers(self, db, sql_ds, minio_ds):
        mod = _make_module(db, sql_ds, minio_ds)
        tpl = _make_template(db, mod)
        m = _make_mapping(db, tpl, write_headers=True, sql_content="SELECT 1 AS id, 'Alice' AS name")
        exc = _make_execution(db, tpl)

        result, output = _run_executor(db, mod, tpl, [m], exc)
        assert result.status == ExecutionStatusEnum.SUCCESS

        wb = _load_output(output)
        ws = wb["Sheet1"]
        assert ws["A1"].value == "id"
        assert ws["B1"].value == "name"
        assert ws["A2"].value == 1
        assert ws["B2"].value == "Alice"
        wb.close()
        os.unlink(output)


# ---------------------------------------------------------------------------
# 2. SINGLE mode
# ---------------------------------------------------------------------------


class TestSingleMode:
    def test_single_value(self, db, sql_ds, minio_ds):
        mod = _make_module(db, sql_ds, minio_ds)
        tpl = _make_template(db, mod)
        m = _make_mapping(db, tpl, write_mode=SheetWriteModeEnum.SINGLE,
                          start_cell="C5", sql_content="SELECT 42 AS total")
        exc = _make_execution(db, tpl)

        result, output = _run_executor(db, mod, tpl, [m], exc)
        assert result.status == ExecutionStatusEnum.SUCCESS

        wb = _load_output(output)
        ws = wb["Sheet1"]
        assert ws["C5"].value == 42
        assert ws["A1"].value is None
        wb.close()
        os.unlink(output)


# ---------------------------------------------------------------------------
# 3. Collision auto-shift + gap_rows
# ---------------------------------------------------------------------------


class TestCollisionAutoShift:
    def test_two_mappings_same_sheet_auto_shift(self, db, sql_ds, minio_ds):
        """Second mapping start_cell overlaps first → auto-shifted."""
        mod = _make_module(db, sql_ds, minio_ds)
        tpl = _make_template(db, mod)
        m1 = _make_mapping(db, tpl, sort_order=0, write_headers=True,
                           sql_content="SELECT 1 AS a UNION ALL SELECT 2 UNION ALL SELECT 3")
        m2 = _make_mapping(db, tpl, sort_order=1, write_headers=True,
                           sql_content="SELECT 'x' AS b UNION ALL SELECT 'y'")
        exc = _make_execution(db, tpl)

        result, output = _run_executor(db, mod, tpl, [m1, m2], exc)
        assert result.status == ExecutionStatusEnum.SUCCESS

        wb = _load_output(output)
        ws = wb["Sheet1"]
        # m1: header at A1, data A2-A4 (3 rows + header = 4 rows, last_row=4)
        assert ws["A1"].value == "a"
        assert ws["A4"].value == 3
        # m2: would start A1, shifted to A5 (last_row=4, gap=0, so 4+0+1=5)
        assert ws["A5"].value == "b"  # header
        assert ws["A6"].value == "x"
        assert ws["A7"].value == "y"
        wb.close()
        os.unlink(output)

    def test_gap_rows_between_mappings(self, db, sql_ds, minio_ds):
        mod = _make_module(db, sql_ds, minio_ds)
        tpl = _make_template(db, mod)
        m1 = _make_mapping(db, tpl, sort_order=0, write_headers=True,
                           sql_content="SELECT 1 AS v")
        # gap_rows=3 on second mapping
        m2 = _make_mapping(db, tpl, sort_order=1, write_headers=True, gap_rows=3,
                           sql_content="SELECT 2 AS v")
        exc = _make_execution(db, tpl)

        result, output = _run_executor(db, mod, tpl, [m1, m2], exc)
        assert result.status == ExecutionStatusEnum.SUCCESS

        wb = _load_output(output)
        ws = wb["Sheet1"]
        # m1: A1=header, A2=data → last_row=2
        assert ws["A1"].value == "v"
        assert ws["A2"].value == 1
        # m2: shifted to row 2+3+1=6
        assert ws["A3"].value is None
        assert ws["A4"].value is None
        assert ws["A5"].value is None
        assert ws["A6"].value == "v"  # header
        assert ws["A7"].value == 2    # data
        wb.close()
        os.unlink(output)

    def test_no_collision_different_sheets(self, db, sql_ds, minio_ds):
        mod = _make_module(db, sql_ds, minio_ds)
        tpl = _make_template(db, mod)
        m1 = _make_mapping(db, tpl, sort_order=0, sheet_name="S1", sql_content="SELECT 1 AS a")
        m2 = _make_mapping(db, tpl, sort_order=1, sheet_name="S2", sql_content="SELECT 2 AS b")
        exc = _make_execution(db, tpl)

        result, output = _run_executor(db, mod, tpl, [m1, m2], exc)
        assert result.status == ExecutionStatusEnum.SUCCESS

        wb = _load_output(output)
        assert wb["S1"]["A1"].value == 1
        assert wb["S2"]["A1"].value == 2
        wb.close()
        os.unlink(output)

    def test_no_shift_when_start_cell_below_last_row(self, db, sql_ds, minio_ds):
        """Second mapping starts below first → no shift needed."""
        mod = _make_module(db, sql_ds, minio_ds)
        tpl = _make_template(db, mod)
        m1 = _make_mapping(db, tpl, sort_order=0, sql_content="SELECT 1 AS a")
        m2 = _make_mapping(db, tpl, sort_order=1, start_cell="A10", sql_content="SELECT 2 AS b")
        exc = _make_execution(db, tpl)

        result, output = _run_executor(db, mod, tpl, [m1, m2], exc)
        assert result.status == ExecutionStatusEnum.SUCCESS

        wb = _load_output(output)
        ws = wb["Sheet1"]
        assert ws["A1"].value == 1
        assert ws["A10"].value == 2
        wb.close()
        os.unlink(output)


# ---------------------------------------------------------------------------
# 4. Format config
# ---------------------------------------------------------------------------


class TestFormatConfig:
    def test_header_bold_and_data_number_format(self, db, sql_ds, minio_ds):
        mod = _make_module(db, sql_ds, minio_ds)
        tpl = _make_template(db, mod)
        m = _make_mapping(db, tpl, write_headers=True,
                          sql_content="SELECT 1234.5 AS amount",
                          format_config={
                              "header": {"font": {"bold": True, "size": 12}},
                              "data": {"number_format": "#,##0.00"},
                          })
        exc = _make_execution(db, tpl)

        result, output = _run_executor(db, mod, tpl, [m], exc)
        assert result.status == ExecutionStatusEnum.SUCCESS

        wb = _load_output(output)
        ws = wb["Sheet1"]
        assert ws["A1"].font.bold is True
        assert ws["A1"].font.size == 12
        assert ws["A2"].number_format == "#,##0.00"
        wb.close()
        os.unlink(output)

    def test_fill_and_border(self, db, sql_ds, minio_ds):
        mod = _make_module(db, sql_ds, minio_ds)
        tpl = _make_template(db, mod)
        m = _make_mapping(db, tpl, write_headers=True,
                          sql_content="SELECT 1 AS v",
                          format_config={
                              "header": {
                                  "fill": {"bg_color": "FFFF00"},
                                  "border": {"style": "thin", "color": "000000"},
                              },
                          })
        exc = _make_execution(db, tpl)

        result, output = _run_executor(db, mod, tpl, [m], exc)
        assert result.status == ExecutionStatusEnum.SUCCESS

        wb = _load_output(output)
        ws = wb["Sheet1"]
        assert ws["A1"].fill.fill_type == "solid"
        assert ws["A1"].border.left.style == "thin"
        wb.close()
        os.unlink(output)

    def test_column_widths(self, db, sql_ds, minio_ds):
        mod = _make_module(db, sql_ds, minio_ds)
        tpl = _make_template(db, mod)
        m = _make_mapping(db, tpl, sql_content="SELECT 1 AS a, 2 AS b",
                          format_config={"column_widths": {"A": 20, "B": 30}})
        exc = _make_execution(db, tpl)

        result, output = _run_executor(db, mod, tpl, [m], exc)
        assert result.status == ExecutionStatusEnum.SUCCESS

        wb = _load_output(output)
        ws = wb["Sheet1"]
        assert ws.column_dimensions["A"].width == 20
        assert ws.column_dimensions["B"].width == 30
        wb.close()
        os.unlink(output)

    def test_single_mode_applies_data_format(self, db, sql_ds, minio_ds):
        mod = _make_module(db, sql_ds, minio_ds)
        tpl = _make_template(db, mod)
        m = _make_mapping(db, tpl, write_mode=SheetWriteModeEnum.SINGLE,
                          sql_content="SELECT 42 AS v",
                          format_config={"data": {"font": {"italic": True}}})
        exc = _make_execution(db, tpl)

        result, output = _run_executor(db, mod, tpl, [m], exc)
        assert result.status == ExecutionStatusEnum.SUCCESS

        wb = _load_output(output)
        ws = wb["Sheet1"]
        assert ws["A1"].font.italic is True
        wb.close()
        os.unlink(output)


# ---------------------------------------------------------------------------
# 5. Auto-fit + Wrap text
# ---------------------------------------------------------------------------


class TestAutoFitWrapText:
    def test_auto_fit_sets_column_widths(self, db, sql_ds, minio_ds):
        mod = _make_module(db, sql_ds, minio_ds)
        tpl = _make_template(db, mod)
        m = _make_mapping(db, tpl, write_headers=True,
                          sql_content="SELECT 'short' AS col_a, 'a much longer text value here' AS col_b",
                          format_config={"auto_fit": True, "auto_fit_max_width": 25})
        exc = _make_execution(db, tpl)

        result, output = _run_executor(db, mod, tpl, [m], exc)
        assert result.status == ExecutionStatusEnum.SUCCESS

        wb = _load_output(output)
        ws = wb["Sheet1"]
        w_a = ws.column_dimensions["A"].width
        w_b = ws.column_dimensions["B"].width
        assert w_a < w_b
        assert w_b <= 27  # max 25 + 2
        wb.close()
        os.unlink(output)

    def test_auto_fit_skipped_when_column_widths_set(self, db, sql_ds, minio_ds):
        mod = _make_module(db, sql_ds, minio_ds)
        tpl = _make_template(db, mod)
        m = _make_mapping(db, tpl, write_headers=True,
                          sql_content="SELECT 'long value here' AS x",
                          format_config={"auto_fit": True, "column_widths": {"A": 10}})
        exc = _make_execution(db, tpl)

        result, output = _run_executor(db, mod, tpl, [m], exc)
        assert result.status == ExecutionStatusEnum.SUCCESS

        wb = _load_output(output)
        ws = wb["Sheet1"]
        assert ws.column_dimensions["A"].width == 10
        wb.close()
        os.unlink(output)

    def test_wrap_text_applied_to_all_cells(self, db, sql_ds, minio_ds):
        mod = _make_module(db, sql_ds, minio_ds)
        tpl = _make_template(db, mod)
        m = _make_mapping(db, tpl, write_headers=True,
                          sql_content="SELECT 1 AS v",
                          format_config={"wrap_text": True})
        exc = _make_execution(db, tpl)

        result, output = _run_executor(db, mod, tpl, [m], exc)
        assert result.status == ExecutionStatusEnum.SUCCESS

        wb = _load_output(output)
        ws = wb["Sheet1"]
        assert ws["A1"].alignment.wrap_text is True  # header
        assert ws["A2"].alignment.wrap_text is True  # data
        wb.close()
        os.unlink(output)


# ---------------------------------------------------------------------------
# 6. Template-level format merged with mapping override
# ---------------------------------------------------------------------------


class TestFormatMerge:
    def test_template_format_applies_to_all_mappings(self, db, sql_ds, minio_ds):
        mod = _make_module(db, sql_ds, minio_ds)
        tpl = _make_template(db, mod, format_config={
            "header": {"font": {"bold": True}},
            "data": {"number_format": "#,##0"},
        })
        m = _make_mapping(db, tpl, write_headers=True, sql_content="SELECT 1000 AS val")
        exc = _make_execution(db, tpl)

        result, output = _run_executor(db, mod, tpl, [m], exc)
        assert result.status == ExecutionStatusEnum.SUCCESS

        wb = _load_output(output)
        ws = wb["Sheet1"]
        assert ws["A1"].font.bold is True
        assert ws["A2"].number_format == "#,##0"
        wb.close()
        os.unlink(output)

    def test_mapping_override_template_format(self, db, sql_ds, minio_ds):
        mod = _make_module(db, sql_ds, minio_ds)
        tpl = _make_template(db, mod, format_config={
            "header": {"font": {"bold": True, "size": 10}},
        })
        m = _make_mapping(db, tpl, write_headers=True,
                          sql_content="SELECT 1 AS v",
                          format_config={"header": {"font": {"size": 16, "italic": True}}})
        exc = _make_execution(db, tpl)

        result, output = _run_executor(db, mod, tpl, [m], exc)
        assert result.status == ExecutionStatusEnum.SUCCESS

        wb = _load_output(output)
        ws = wb["Sheet1"]
        assert ws["A1"].font.bold is True   # from template
        assert ws["A1"].font.size == 16     # overridden by mapping
        assert ws["A1"].font.italic is True # from mapping
        wb.close()
        os.unlink(output)


# ---------------------------------------------------------------------------
# 7. Empty result + inactive mapping
# ---------------------------------------------------------------------------


class TestEdgeCases:
    def test_empty_result_set(self, db, sql_ds, minio_ds):
        mod = _make_module(db, sql_ds, minio_ds)
        tpl = _make_template(db, mod)
        m = _make_mapping(db, tpl, write_headers=True,
                          sql_content="SELECT 1 AS v WHERE false")
        exc = _make_execution(db, tpl)

        result, output = _run_executor(db, mod, tpl, [m], exc)
        assert result.status == ExecutionStatusEnum.SUCCESS

        wb = _load_output(output)
        ws = wb["Sheet1"]
        assert ws["A1"].value is None
        wb.close()
        os.unlink(output)

    def test_inactive_mapping_skipped(self, db, sql_ds, minio_ds):
        mod = _make_module(db, sql_ds, minio_ds)
        tpl = _make_template(db, mod)
        m1 = _make_mapping(db, tpl, sort_order=0, sql_content="SELECT 1 AS a")
        m2 = _make_mapping(db, tpl, sort_order=1, is_active=False,
                           sql_content="SELECT 999 AS should_not_appear")
        exc = _make_execution(db, tpl)

        result, output = _run_executor(db, mod, tpl, [m1, m2], exc)
        assert result.status == ExecutionStatusEnum.SUCCESS

        wb = _load_output(output)
        ws = wb["Sheet1"]
        assert ws["A1"].value == 1
        # m2 would write to A2 or shift, but it's inactive
        assert ws["A2"].value is None
        wb.close()
        os.unlink(output)

    def test_multiple_mappings_mixed_modes(self, db, sql_ds, minio_ds):
        """SINGLE + ROWS on same sheet."""
        mod = _make_module(db, sql_ds, minio_ds)
        tpl = _make_template(db, mod)
        m_single = _make_mapping(db, tpl, sort_order=0, start_cell="D1",
                                 write_mode=SheetWriteModeEnum.SINGLE,
                                 sql_content="SELECT 'Total' AS label")
        m_rows = _make_mapping(db, tpl, sort_order=1, start_cell="A1", write_headers=True,
                               sql_content="SELECT 1 AS id, 'x' AS val UNION ALL SELECT 2, 'y'")
        exc = _make_execution(db, tpl)

        result, output = _run_executor(db, mod, tpl, [m_single, m_rows], exc)
        assert result.status == ExecutionStatusEnum.SUCCESS

        wb = _load_output(output)
        ws = wb["Sheet1"]
        assert ws["D1"].value == "Total"
        # SINGLE at D1 sets last_row=1 → ROWS at A1 shifts to A2
        assert ws["A2"].value == "id"
        assert ws["A3"].value == 1
        wb.close()
        os.unlink(output)


# ---------------------------------------------------------------------------
# 8. Chunked pagination (auto-paginate)
# ---------------------------------------------------------------------------


class TestChunkedPagination:
    def test_auto_paginate_large_result(self, db, sql_ds, minio_ds):
        """Generate >chunk_size rows via generate_series, verify all written."""
        mod = _make_module(db, sql_ds, minio_ds)
        tpl = _make_template(db, mod)
        m = _make_mapping(db, tpl, write_headers=True,
                          sql_content="SELECT i AS num FROM generate_series(1, 100) AS i")
        exc = _make_execution(db, tpl)

        with patch.object(_settings, "REPORT_SQL_CHUNK_SIZE", 30):
            result, output = _run_executor(db, mod, tpl, [m], exc)

        assert result.status == ExecutionStatusEnum.SUCCESS

        wb = _load_output(output)
        ws = wb["Sheet1"]
        assert ws["A1"].value == "num"
        assert ws["A2"].value == 1
        assert ws["A101"].value == 100
        assert ws["A102"].value is None
        wb.close()
        os.unlink(output)

    def test_auto_paginate_with_format(self, db, sql_ds, minio_ds):
        mod = _make_module(db, sql_ds, minio_ds)
        tpl = _make_template(db, mod)
        m = _make_mapping(db, tpl, write_headers=True,
                          sql_content="SELECT i AS num FROM generate_series(1, 10) AS i",
                          format_config={
                              "header": {"font": {"bold": True}},
                              "data": {"number_format": "0.00"},
                              "wrap_text": True,
                          })
        exc = _make_execution(db, tpl)

        with patch.object(_settings, "REPORT_SQL_CHUNK_SIZE", 3):
            result, output = _run_executor(db, mod, tpl, [m], exc)

        assert result.status == ExecutionStatusEnum.SUCCESS

        wb = _load_output(output)
        ws = wb["Sheet1"]
        assert ws["A1"].font.bold is True
        assert ws["A1"].alignment.wrap_text is True
        assert ws["A2"].number_format == "0.00"
        assert ws["A2"].alignment.wrap_text is True
        assert ws["A11"].value == 10
        wb.close()
        os.unlink(output)


# ---------------------------------------------------------------------------
# 9. Collision + format combined scenario
# ---------------------------------------------------------------------------


class TestCombinedScenario:
    def test_full_scenario(self, db, sql_ds, minio_ds):
        """Multi-mapping: collision + gap + format + auto-fit all together."""
        mod = _make_module(db, sql_ds, minio_ds)
        tpl = _make_template(db, mod, format_config={
            "header": {"font": {"bold": True}},
        })
        m1 = _make_mapping(db, tpl, sort_order=0, write_headers=True,
                           sql_content="SELECT 'hello world' AS message, 12345 AS amount",
                           format_config={"auto_fit": True, "auto_fit_max_width": 20})
        m2 = _make_mapping(db, tpl, sort_order=1, write_headers=True, gap_rows=2,
                           sql_content="SELECT 'summary' AS label, 99999 AS total",
                           format_config={
                               "header": {"fill": {"bg_color": "FFFF00"}},
                               "data": {"number_format": "#,##0"},
                           })
        exc = _make_execution(db, tpl)

        result, output = _run_executor(db, mod, tpl, [m1, m2], exc)
        assert result.status == ExecutionStatusEnum.SUCCESS

        wb = _load_output(output)
        ws = wb["Sheet1"]

        # m1: A1=header, A2=data → last_row=2
        assert ws["A1"].value == "message"
        assert ws["A1"].font.bold is True  # template format
        assert ws["B2"].value == 12345

        # auto-fit applied
        assert ws.column_dimensions["A"].width > 8

        # m2: shifted to 2+2+1=5, with gap
        assert ws["A3"].value is None  # gap
        assert ws["A4"].value is None  # gap
        assert ws["A5"].value == "label"
        assert ws["A5"].font.bold is True       # template format
        assert ws["A5"].fill.fill_type == "solid" # mapping override
        assert ws["B6"].value == 99999
        assert ws["B6"].number_format == "#,##0"

        wb.close()
        os.unlink(output)
