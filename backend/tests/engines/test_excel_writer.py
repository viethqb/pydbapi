"""Unit tests for excel_writer format + executor deep merge / collision shift."""
from openpyxl import Workbook

from app.engines.excel.excel_writer import (
    _apply_cell_format,
    _col_letter,
    _parse_cell,
    write_rows,
    write_single,
)
from app.engines.excel.executor import _deep_merge, _effective_start


def _make_wb(sheet="S") -> Workbook:
    wb = Workbook()
    wb.active.title = sheet
    return wb


# ---------------- helpers ----------------


def test_parse_and_col_letter_roundtrip():
    assert _parse_cell("A1") == (1, 1)
    assert _parse_cell("B5") == (5, 2)
    assert _parse_cell("AA10") == (10, 27)
    assert _col_letter(1) == "A"
    assert _col_letter(27) == "AA"


# ---------------- format apply ----------------


def test_apply_header_format_sets_font_fill_border_alignment_and_number_format():
    wb = _make_wb()
    ws = wb["S"]
    cell = ws.cell(row=1, column=1, value="x")
    fmt = {
        "font": {"bold": True, "color": "FFFF0000", "size": 14},
        "fill": {"bg_color": "FFFFFF00"},
        "border": {"style": "thin", "color": "FF000000"},
        "alignment": {"horizontal": "center", "wrap_text": True},
        "number_format": "#,##0.00",
    }
    _apply_cell_format(cell, fmt)
    assert cell.font.bold is True
    assert cell.font.size == 14
    assert cell.fill.fill_type == "solid"
    assert cell.border.left.style == "thin"
    assert cell.alignment.horizontal == "center"
    assert cell.alignment.wrap_text is True
    assert cell.number_format == "#,##0.00"


def test_write_rows_applies_header_and_data_formats():
    wb = _make_wb()
    data = [{"a": 1, "b": 2}, {"a": 3, "b": 4}]
    write_rows(
        wb, "S", "A1", data,
        write_headers=True,
        header_format={"font": {"bold": True}},
        data_format={"number_format": "0.00"},
        column_widths={"A": 12.5},
    )
    ws = wb["S"]
    assert ws["A1"].value == "a"
    assert ws["A1"].font.bold is True
    assert ws["A2"].value == 1
    assert ws["A2"].number_format == "0.00"
    assert ws["A3"].value == 3
    assert ws.column_dimensions["A"].width == 12.5


def test_write_single_applies_data_format():
    wb = _make_wb()
    write_single(wb, "S", "B2", [{"v": 99}], data_format={"font": {"italic": True}})
    cell = wb["S"]["B2"]
    assert cell.value == 99
    assert cell.font.italic is True


def test_write_rows_empty_is_noop():
    wb = _make_wb()
    assert write_rows(wb, "S", "A1", []) == 0


# ---------------- executor helpers ----------------


def test_deep_merge_override_wins_per_key():
    base = {"header": {"font": {"bold": True, "size": 12}}, "data": {"number_format": "0"}}
    override = {"header": {"font": {"size": 14, "color": "FF0000"}}}
    out = _deep_merge(base, override)
    assert out["header"]["font"] == {"bold": True, "size": 14, "color": "FF0000"}
    assert out["data"] == {"number_format": "0"}


def test_deep_merge_handles_none():
    assert _deep_merge(None, None) is None
    assert _deep_merge({"a": 1}, None) == {"a": 1}
    assert _deep_merge(None, {"a": 1}) == {"a": 1}


def test_effective_start_no_collision_returns_original():
    last: dict[str, int] = {}
    cell, shifted = _effective_start("S", "A5", 0, last)
    assert cell == "A5"
    assert shifted is False


def test_effective_start_collision_shifts_by_gap():
    last = {"S": 10}
    cell, shifted = _effective_start("S", "A1", 2, last)
    assert shifted is True
    assert cell == "A13"  # 10 + 2 (gap) + 1


def test_effective_start_collision_no_gap():
    last = {"S": 10}
    cell, shifted = _effective_start("S", "B5", 0, last)
    assert shifted is True
    assert cell == "B11"  # column preserved


def test_effective_start_cross_sheet_independent():
    last = {"S1": 100}
    cell, shifted = _effective_start("S2", "A1", 5, last)
    assert shifted is False
    assert cell == "A1"
