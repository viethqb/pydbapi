"""Unit tests for the report engine components."""

import os
import tempfile
import uuid
from datetime import date, datetime, time
from decimal import Decimal

import pytest
from openpyxl import Workbook, load_workbook


# ---------------------------------------------------------------------------
# excel_writer tests
# ---------------------------------------------------------------------------


class TestParseCell:
    def test_simple(self):
        from app.engines.excel.excel_writer import _parse_cell
        assert _parse_cell("A1") == (1, 1)
        assert _parse_cell("B5") == (5, 2)
        assert _parse_cell("Z1") == (1, 26)

    def test_multi_letter(self):
        from app.engines.excel.excel_writer import _parse_cell
        assert _parse_cell("AA1") == (1, 27)
        assert _parse_cell("AZ1") == (1, 52)
        assert _parse_cell("BA1") == (1, 53)

    def test_lowercase(self):
        from app.engines.excel.excel_writer import _parse_cell
        assert _parse_cell("a1") == (1, 1)
        assert _parse_cell("bc10") == (10, 55)

    def test_invalid(self):
        from app.engines.excel.excel_writer import _parse_cell
        with pytest.raises(ValueError):
            _parse_cell("")
        with pytest.raises(ValueError):
            _parse_cell("123")
        with pytest.raises(ValueError):
            _parse_cell("A")


class TestColLetter:
    def test_basic(self):
        from app.engines.excel.excel_writer import _col_letter
        assert _col_letter(1) == "A"
        assert _col_letter(2) == "B"
        assert _col_letter(26) == "Z"

    def test_multi_letter(self):
        from app.engines.excel.excel_writer import _col_letter
        assert _col_letter(27) == "AA"
        assert _col_letter(28) == "AB"
        assert _col_letter(52) == "AZ"
        assert _col_letter(53) == "BA"

    def test_roundtrip(self):
        from app.engines.excel.excel_writer import _col_letter, _parse_cell
        for col in [1, 5, 26, 27, 52, 100, 256]:
            letter = _col_letter(col)
            row, parsed_col = _parse_cell(f"{letter}1")
            assert parsed_col == col, f"Roundtrip failed for col {col}: letter={letter}, parsed={parsed_col}"


class TestSanitizeValue:
    def test_primitives(self):
        from app.engines.excel.excel_writer import _sanitize_value
        assert _sanitize_value(None) is None
        assert _sanitize_value("hello") == "hello"
        assert _sanitize_value(42) == 42
        assert _sanitize_value(3.14) == 3.14
        assert _sanitize_value(True) is True

    def test_uuid(self):
        from app.engines.excel.excel_writer import _sanitize_value
        u = uuid.uuid4()
        assert _sanitize_value(u) == str(u)

    def test_decimal(self):
        from app.engines.excel.excel_writer import _sanitize_value
        assert _sanitize_value(Decimal("10.5")) == 10.5

    def test_datetime(self):
        from app.engines.excel.excel_writer import _sanitize_value
        dt = datetime(2026, 1, 1, 12, 0)
        assert _sanitize_value(dt) == dt  # openpyxl handles datetime

    def test_bytes(self):
        from app.engines.excel.excel_writer import _sanitize_value
        assert _sanitize_value(b"\x00\xff") == "00ff"

    def test_fallback(self):
        from app.engines.excel.excel_writer import _sanitize_value
        assert _sanitize_value({"key": "val"}) == "{'key': 'val'}"


class TestWriteRows:
    def test_basic(self):
        from app.engines.excel.excel_writer import write_rows
        wb = Workbook()
        ws = wb.active
        ws.title = "Test"
        data = [
            {"id": 1, "name": "Alice"},
            {"id": 2, "name": "Bob"},
        ]
        rows_written = write_rows(wb, "Test", "A1", data)
        assert rows_written == 2
        assert ws["A1"].value == 1
        assert ws["B1"].value == "Alice"
        assert ws["A2"].value == 2

    def test_with_headers(self):
        from app.engines.excel.excel_writer import write_rows
        wb = Workbook()
        ws = wb.active
        ws.title = "Test"
        data = [{"id": 1, "name": "Alice"}]
        rows_written = write_rows(wb, "Test", "B3", data, write_headers=True)
        assert rows_written == 2  # header + 1 data row
        assert ws["B3"].value == "id"
        assert ws["C3"].value == "name"
        assert ws["B4"].value == 1

    def test_empty_data(self):
        from app.engines.excel.excel_writer import write_rows
        wb = Workbook()
        ws = wb.active
        ws.title = "Test"
        assert write_rows(wb, "Test", "A1", []) == 0

    def test_uuid_sanitized(self):
        from app.engines.excel.excel_writer import write_rows
        wb = Workbook()
        ws = wb.active
        ws.title = "Test"
        u = uuid.uuid4()
        write_rows(wb, "Test", "A1", [{"id": u}])
        assert ws["A1"].value == str(u)

    def test_sheet_not_found(self):
        from app.engines.excel.excel_writer import write_rows
        wb = Workbook()
        with pytest.raises(ValueError, match="not found"):
            write_rows(wb, "NonExistent", "A1", [{"id": 1}])


class TestWriteSingle:
    def test_basic(self):
        from app.engines.excel.excel_writer import write_single
        wb = Workbook()
        ws = wb.active
        ws.title = "Test"
        write_single(wb, "Test", "C5", [{"count": 42}])
        assert ws["C5"].value == 42

    def test_empty(self):
        from app.engines.excel.excel_writer import write_single
        wb = Workbook()
        ws = wb.active
        ws.title = "Test"
        write_single(wb, "Test", "A1", [])  # No error


# ---------------------------------------------------------------------------
# excel_extractor tests
# ---------------------------------------------------------------------------


class TestExtractSheet:
    def test_extract(self):
        from app.engines.excel.excel_extractor import extract_sheet

        # Create source workbook with 2 sheets
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Data"
        ws1["A1"] = "raw data"
        ws2 = wb.create_sheet("Summary")
        ws2["A1"] = "Total"
        ws2["B1"] = 42

        src = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        dst = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        try:
            wb.save(src.name)
            wb.close()
            extract_sheet(src.name, "Summary", dst.name)

            result = load_workbook(dst.name)
            assert result.sheetnames == ["Summary"]
            assert result["Summary"]["A1"].value == "Total"
            assert result["Summary"]["B1"].value == 42
            result.close()
        finally:
            os.unlink(src.name)
            os.unlink(dst.name)

    def test_sheet_not_found(self):
        from app.engines.excel.excel_extractor import extract_sheet

        wb = Workbook()
        src = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        try:
            wb.save(src.name)
            wb.close()
            with pytest.raises(ValueError, match="not found"):
                extract_sheet(src.name, "NonExistent", "/tmp/out.xlsx")
        finally:
            os.unlink(src.name)


# ---------------------------------------------------------------------------
# validate_path tests
# ---------------------------------------------------------------------------


class TestValidatePath:
    def test_valid(self):
        from app.api.routes.report_modules import _validate_path
        assert _validate_path("templates/file.xlsx") == "templates/file.xlsx"
        assert _validate_path("my-bucket") == "my-bucket"

    def test_traversal(self):
        from app.api.routes.report_modules import _validate_path
        with pytest.raises(Exception):
            _validate_path("../etc/passwd")
        with pytest.raises(Exception):
            _validate_path("bucket/../secret")

    def test_absolute(self):
        from app.api.routes.report_modules import _validate_path
        with pytest.raises(Exception):
            _validate_path("/etc/passwd")

    def test_dash_prefix(self):
        from app.api.routes.report_modules import _validate_path
        with pytest.raises(Exception):
            _validate_path("--malicious")
