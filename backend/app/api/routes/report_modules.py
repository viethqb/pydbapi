"""Report Module CRUD, Template CRUD, Mapping CRUD, Client Access, Generate."""
import logging
import uuid
from datetime import UTC, datetime
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Request
from sqlmodel import Session, func, select

from app.api.deps import (
    CurrentUser,
    SessionDep,
    require_permission,
    require_permission_for_body_resource,
    require_permission_for_resource,
)
from app.api.pagination import get_allowed_ids, paginate
from app.core.config import settings
from app.core.permission_resources import (
    ensure_resource_permissions,
    remove_resource_permissions,
)
from app.engines.excel.executor import ExcelReportExecutor
from app.models import Message, User
from app.models_permission import PermissionActionEnum, ResourceTypeEnum
from app.models_report import (
    ExecutionStatusEnum,
    ReportExecution,
    ReportModule,
    ReportModuleClientLink,
    ReportSheetMapping,
    ReportTemplate,
    ReportTemplateClientLink,
)
from app.schemas_report import (
    ClientIdsOut,
    ClientIdsUpdate,
    ReportExecutionListOut,
    ReportExecutionPublic,
    ReportGenerateIn,
    ReportGenerateOut,
    ReportModuleCreate,
    ReportModuleDetail,
    ReportModuleListIn,
    ReportModuleListOut,
    ReportModulePublic,
    ReportModuleUpdate,
    ReportTemplateCreate,
    ReportTemplateDetail,
    ReportTemplateListIn,
    ReportTemplateListOut,
    ReportTemplatePublic,
    ReportTemplateUpdate,
    SheetMappingBatchUpdate,
    SheetMappingCreate,
    SheetMappingPublic,
    SheetMappingUpdate,
)

_log = logging.getLogger(__name__)
router = APIRouter(prefix="/report-modules", tags=["report-modules"])
RES = ResourceTypeEnum.REPORT_MODULE
ACTIONS = (PermissionActionEnum.READ, PermissionActionEnum.CREATE, PermissionActionEnum.UPDATE, PermissionActionEnum.DELETE, PermissionActionEnum.EXECUTE)


def _mid_from_path(*, id: uuid.UUID, **_: Any) -> uuid.UUID | None:
    return id

def _mid_from_body(*, body: Any, **_: Any) -> uuid.UUID | None:
    return getattr(body, "id", None)


# ========================= Converters =========================

def _module_public(m: ReportModule) -> ReportModulePublic:
    return ReportModulePublic(**{k: getattr(m, k) for k in ReportModulePublic.model_fields})

def _module_detail(m: ReportModule) -> ReportModuleDetail:
    base = _module_public(m).model_dump()
    templates = [_template_public(t) for t in sorted(m.templates or [], key=lambda t: t.name)]
    client_ids = [l.app_client_id for l in (m.client_links or [])]
    return ReportModuleDetail(**base, templates=templates, client_ids=client_ids)

def _template_public(t: ReportTemplate) -> ReportTemplatePublic:
    return ReportTemplatePublic(**{k: getattr(t, k) for k in ReportTemplatePublic.model_fields})

def _template_detail(t: ReportTemplate) -> ReportTemplateDetail:
    base = _template_public(t).model_dump()
    mappings = [_mapping_public(m) for m in sorted(t.sheet_mappings or [], key=lambda m: m.sort_order)]
    client_ids = [l.app_client_id for l in (t.client_links or [])]
    return ReportTemplateDetail(**base, sheet_mappings=mappings, client_ids=client_ids)

def _mapping_public(m: ReportSheetMapping) -> SheetMappingPublic:
    return SheetMappingPublic(**{k: getattr(m, k) for k in SheetMappingPublic.model_fields})

def _execution_public(e: ReportExecution) -> ReportExecutionPublic:
    return ReportExecutionPublic(**{k: getattr(e, k) for k in ReportExecutionPublic.model_fields})


# ========================= MinIO Buckets =========================


@router.get("/buckets/{datasource_id}", response_model=list[str],
    dependencies=[Depends(require_permission(RES, PermissionActionEnum.READ))])
def list_minio_buckets(session: SessionDep, datasource_id: uuid.UUID) -> Any:
    """List all buckets from a MinIO datasource."""
    from app.engines.excel.minio_client import get_minio_client
    from app.models_dbapi import DataSource

    ds = session.get(DataSource, datasource_id)
    if not ds:
        raise HTTPException(404, "Datasource not found")
    try:
        client = get_minio_client(ds)
        buckets = client.list_buckets()
        return sorted([b.name for b in buckets])
    except Exception as e:
        raise HTTPException(400, f"Failed to list buckets: {e}")


def _validate_path(path: str) -> str:
    """Reject path traversal attempts."""
    if ".." in path or path.startswith("/") or path.startswith("-"):
        raise HTTPException(400, "Invalid path")
    return path


@router.get("/files/{datasource_id}/{bucket}", response_model=list[str],
    dependencies=[Depends(require_permission(RES, PermissionActionEnum.READ))])
def list_minio_files(session: SessionDep, datasource_id: uuid.UUID, bucket: str, prefix: str = "", suffix: str = ".xlsx") -> Any:
    """List files in a MinIO bucket, optionally filtered by prefix and suffix."""
    from app.engines.excel.minio_client import get_minio_client
    from app.models_dbapi import DataSource

    _validate_path(bucket)
    if prefix:
        _validate_path(prefix)
    ds = session.get(DataSource, datasource_id)
    if not ds:
        raise HTTPException(404, "Datasource not found")
    try:
        client = get_minio_client(ds)
        objects = client.list_objects(bucket, prefix=prefix or None, recursive=True)
        files = [obj.object_name for obj in objects if not obj.is_dir and obj.object_name.endswith(suffix)]
        return sorted(files)
    except Exception as e:
        _log.warning("Failed to list files: %s", e)
        raise HTTPException(400, "Unable to list files")


@router.get("/sheets/{datasource_id}/{bucket}/{file_path:path}", response_model=list[str],
    dependencies=[Depends(require_permission(RES, PermissionActionEnum.READ))])
def list_excel_sheets(session: SessionDep, datasource_id: uuid.UUID, bucket: str, file_path: str) -> Any:
    """Download an xlsx file from MinIO and return its sheet names."""
    import os
    import tempfile

    from openpyxl import load_workbook

    from app.engines.excel.minio_client import download_file, get_minio_client
    from app.models_dbapi import DataSource

    _validate_path(bucket)
    _validate_path(file_path)
    ds = session.get(DataSource, datasource_id)
    if not ds:
        raise HTTPException(404, "Datasource not found")
    tmp_path = None
    try:
        client = get_minio_client(ds)
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp_path = tmp.name
        tmp.close()
        download_file(client, bucket, file_path, tmp_path)
        wb = load_workbook(tmp_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        return sheets
    except HTTPException:
        raise
    except Exception as e:
        _log.warning("Failed to read sheets: %s", e)
        raise HTTPException(400, "Unable to read sheets")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)


# ========================= Module CRUD =========================

def _module_filters(stmt: Any, body: ReportModuleListIn) -> Any:
    if body.name__ilike:
        stmt = stmt.where(ReportModule.name.ilike(f"%{body.name__ilike}%"))
    if body.is_active is not None:
        stmt = stmt.where(ReportModule.is_active == body.is_active)
    return stmt


@router.post("/list", response_model=ReportModuleListOut,
    dependencies=[Depends(require_permission(RES, PermissionActionEnum.READ))])
def list_modules(session: SessionDep, current_user: CurrentUser, body: ReportModuleListIn) -> Any:
    allowed = get_allowed_ids(session, current_user, RES, PermissionActionEnum.READ)
    data, total = paginate(session, ReportModule, body, filters_fn=_module_filters, allowed_ids=allowed, order_by=ReportModule.name, to_public=_module_public)
    return ReportModuleListOut(data=data, total=total)


@router.post("/create", response_model=ReportModuleDetail,
    dependencies=[Depends(require_permission(RES, PermissionActionEnum.CREATE))])
def create_module(session: SessionDep, current_user: CurrentUser, body: ReportModuleCreate) -> Any:
    if session.exec(select(ReportModule).where(ReportModule.name == body.name)).first():
        raise HTTPException(400, f"Module name '{body.name}' already exists")
    mod = ReportModule(**body.model_dump())
    session.add(mod)
    session.flush()
    ensure_resource_permissions(session, RES, mod.id, ACTIONS)
    session.commit()
    session.refresh(mod)
    return _module_detail(mod)


@router.post("/update", response_model=ReportModulePublic)
def update_module(session: SessionDep, body: ReportModuleUpdate,
    _: User = Depends(require_permission_for_body_resource(RES, PermissionActionEnum.UPDATE, ReportModuleUpdate, _mid_from_body))) -> Any:
    mod = session.get(ReportModule, body.id)
    if not mod:
        raise HTTPException(404, "Module not found")
    data = body.model_dump(exclude_unset=True, exclude={"id"})
    if "name" in data and data["name"] != mod.name:
        if session.exec(select(ReportModule).where(ReportModule.name == data["name"], ReportModule.id != body.id)).first():
            raise HTTPException(400, f"Name '{data['name']}' already exists")
    mod.sqlmodel_update(data)
    mod.updated_at = datetime.now(UTC)
    session.add(mod)
    session.commit()
    session.refresh(mod)
    return _module_public(mod)


@router.post("/delete", response_model=Message)
def delete_module(session: SessionDep, id: uuid.UUID,
    _: User = Depends(require_permission_for_resource(RES, PermissionActionEnum.DELETE, _mid_from_path))) -> Any:
    mod = session.get(ReportModule, id)
    if not mod:
        raise HTTPException(404, "Module not found")
    # Explicit child cleanup
    for t in session.exec(select(ReportTemplate).where(ReportTemplate.report_module_id == id)).all():
        for e in session.exec(select(ReportExecution).where(ReportExecution.report_template_id == t.id)).all():
            session.delete(e)
        for m in session.exec(select(ReportSheetMapping).where(ReportSheetMapping.report_template_id == t.id)).all():
            session.delete(m)
        for l in session.exec(select(ReportTemplateClientLink).where(ReportTemplateClientLink.report_template_id == t.id)).all():
            session.delete(l)
        session.delete(t)
    for l in session.exec(select(ReportModuleClientLink).where(ReportModuleClientLink.report_module_id == id)).all():
        session.delete(l)
    remove_resource_permissions(session, RES, id)
    session.delete(mod)
    session.commit()
    return Message(message="Module deleted")


@router.get("/{id}", response_model=ReportModuleDetail)
def get_module(session: SessionDep, id: uuid.UUID,
    _: User = Depends(require_permission_for_resource(RES, PermissionActionEnum.READ, _mid_from_path))) -> Any:
    mod = session.get(ReportModule, id)
    if not mod:
        raise HTTPException(404, "Module not found")
    return _module_detail(mod)


# ========================= Module Clients =========================

@router.get("/{id}/clients", response_model=ClientIdsOut)
def get_module_clients(session: SessionDep, id: uuid.UUID,
    _: User = Depends(require_permission_for_resource(RES, PermissionActionEnum.READ, _mid_from_path))) -> Any:
    mod = session.get(ReportModule, id)
    if not mod:
        raise HTTPException(404, "Module not found")
    return ClientIdsOut(client_ids=[l.app_client_id for l in (mod.client_links or [])])


@router.post("/{id}/clients", response_model=ClientIdsOut)
def set_module_clients(session: SessionDep, id: uuid.UUID, body: ClientIdsUpdate,
    _: User = Depends(require_permission_for_resource(RES, PermissionActionEnum.UPDATE, _mid_from_path))) -> Any:
    mod = session.get(ReportModule, id)
    if not mod:
        raise HTTPException(404, "Module not found")
    for l in session.exec(select(ReportModuleClientLink).where(ReportModuleClientLink.report_module_id == id)).all():
        session.delete(l)
    for cid in body.client_ids:
        session.add(ReportModuleClientLink(report_module_id=id, app_client_id=cid))
    session.commit()
    session.refresh(mod)
    return ClientIdsOut(client_ids=[l.app_client_id for l in (mod.client_links or [])])


# ========================= Template CRUD =========================

def _tpl_filters(stmt: Any, body: ReportTemplateListIn) -> Any:
    if body.name__ilike:
        stmt = stmt.where(ReportTemplate.name.ilike(f"%{body.name__ilike}%"))
    if body.module_id is not None:
        stmt = stmt.where(ReportTemplate.report_module_id == body.module_id)
    if body.is_active is not None:
        stmt = stmt.where(ReportTemplate.is_active == body.is_active)
    return stmt


@router.post("/templates/list", response_model=ReportTemplateListOut,
    dependencies=[Depends(require_permission(RES, PermissionActionEnum.READ))])
def list_all_templates(session: SessionDep, body: ReportTemplateListIn) -> Any:
    """List all templates across all modules, with optional module_id filter."""
    stmt = select(ReportTemplate)
    stmt = _tpl_filters(stmt, body)
    count_stmt = select(func.count()).select_from(ReportTemplate)
    count_stmt = _tpl_filters(count_stmt, body)
    total = session.exec(count_stmt).one()
    offset = (body.page - 1) * body.page_size
    rows = session.exec(stmt.order_by(ReportTemplate.name).offset(offset).limit(body.page_size)).all()
    return ReportTemplateListOut(data=[_template_public(t) for t in rows], total=total)


@router.post("/{id}/templates/list", response_model=ReportTemplateListOut)
def list_templates(session: SessionDep, id: uuid.UUID, body: ReportTemplateListIn,
    _: User = Depends(require_permission_for_resource(RES, PermissionActionEnum.READ, _mid_from_path))) -> Any:
    base = select(ReportTemplate).where(ReportTemplate.report_module_id == id)
    stmt = _tpl_filters(base, body)
    total = session.exec(select(func.count()).select_from(ReportTemplate).where(ReportTemplate.report_module_id == id)).one()
    offset = (body.page - 1) * body.page_size
    rows = session.exec(stmt.order_by(ReportTemplate.name).offset(offset).limit(body.page_size)).all()
    return ReportTemplateListOut(data=[_template_public(t) for t in rows], total=total)


@router.post("/{id}/templates/create", response_model=ReportTemplateDetail)
def create_template(session: SessionDep, id: uuid.UUID, body: ReportTemplateCreate,
    _: User = Depends(require_permission_for_resource(RES, PermissionActionEnum.UPDATE, _mid_from_path))) -> Any:
    mod = session.get(ReportModule, id)
    if not mod:
        raise HTTPException(404, "Module not found")
    if session.exec(select(ReportTemplate).where(ReportTemplate.report_module_id == id, ReportTemplate.name == body.name)).first():
        raise HTTPException(400, f"Template '{body.name}' already exists in this module")
    tpl_data = body.model_dump(exclude={"sheet_mappings"})
    tpl = ReportTemplate(report_module_id=id, **tpl_data)
    session.add(tpl)
    session.flush()
    if body.sheet_mappings:
        for m in body.sheet_mappings:
            session.add(ReportSheetMapping(report_template_id=tpl.id, **m.model_dump()))
    session.commit()
    session.refresh(tpl)
    return _template_detail(tpl)


@router.post("/{id}/templates/update", response_model=ReportTemplatePublic)
def update_template(session: SessionDep, id: uuid.UUID, body: ReportTemplateUpdate,
    _: User = Depends(require_permission_for_resource(RES, PermissionActionEnum.UPDATE, _mid_from_path))) -> Any:
    tpl = session.get(ReportTemplate, body.id)
    if not tpl or tpl.report_module_id != id:
        raise HTTPException(404, "Template not found")
    data = body.model_dump(exclude_unset=True, exclude={"id"})
    tpl.sqlmodel_update(data)
    tpl.updated_at = datetime.now(UTC)
    session.add(tpl)
    session.commit()
    session.refresh(tpl)
    return _template_public(tpl)


@router.post("/{id}/templates/delete", response_model=Message)
def delete_template(session: SessionDep, id: uuid.UUID, tid: uuid.UUID,
    _: User = Depends(require_permission_for_resource(RES, PermissionActionEnum.UPDATE, _mid_from_path))) -> Any:
    tpl = session.get(ReportTemplate, tid)
    if not tpl or tpl.report_module_id != id:
        raise HTTPException(404, "Template not found")
    for e in session.exec(select(ReportExecution).where(ReportExecution.report_template_id == tid)).all():
        session.delete(e)
    for m in session.exec(select(ReportSheetMapping).where(ReportSheetMapping.report_template_id == tid)).all():
        session.delete(m)
    for l in session.exec(select(ReportTemplateClientLink).where(ReportTemplateClientLink.report_template_id == tid)).all():
        session.delete(l)
    session.delete(tpl)
    session.commit()
    return Message(message="Template deleted")


@router.get("/{id}/templates/{tid}", response_model=ReportTemplateDetail)
def get_template(session: SessionDep, id: uuid.UUID, tid: uuid.UUID,
    _: User = Depends(require_permission_for_resource(RES, PermissionActionEnum.READ, _mid_from_path))) -> Any:
    tpl = session.get(ReportTemplate, tid)
    if not tpl or tpl.report_module_id != id:
        raise HTTPException(404, "Template not found")
    return _template_detail(tpl)


# ========================= Template Clients =========================

@router.get("/{id}/templates/{tid}/clients", response_model=ClientIdsOut)
def get_template_clients(session: SessionDep, id: uuid.UUID, tid: uuid.UUID,
    _: User = Depends(require_permission_for_resource(RES, PermissionActionEnum.READ, _mid_from_path))) -> Any:
    tpl = session.get(ReportTemplate, tid)
    if not tpl or tpl.report_module_id != id:
        raise HTTPException(404, "Template not found")
    return ClientIdsOut(client_ids=[l.app_client_id for l in (tpl.client_links or [])])


@router.post("/{id}/templates/{tid}/clients", response_model=ClientIdsOut)
def set_template_clients(session: SessionDep, id: uuid.UUID, tid: uuid.UUID, body: ClientIdsUpdate,
    _: User = Depends(require_permission_for_resource(RES, PermissionActionEnum.UPDATE, _mid_from_path))) -> Any:
    tpl = session.get(ReportTemplate, tid)
    if not tpl or tpl.report_module_id != id:
        raise HTTPException(404, "Template not found")
    for l in session.exec(select(ReportTemplateClientLink).where(ReportTemplateClientLink.report_template_id == tid)).all():
        session.delete(l)
    for cid in body.client_ids:
        session.add(ReportTemplateClientLink(report_template_id=tid, app_client_id=cid))
    session.commit()
    session.refresh(tpl)
    return ClientIdsOut(client_ids=[l.app_client_id for l in (tpl.client_links or [])])


# ========================= Mapping CRUD =========================

@router.post("/{id}/templates/{tid}/mappings/create", response_model=SheetMappingPublic)
def create_mapping(session: SessionDep, id: uuid.UUID, tid: uuid.UUID, body: SheetMappingCreate,
    _: User = Depends(require_permission_for_resource(RES, PermissionActionEnum.UPDATE, _mid_from_path))) -> Any:
    tpl = session.get(ReportTemplate, tid)
    if not tpl or tpl.report_module_id != id:
        raise HTTPException(404, "Template not found")
    mapping = ReportSheetMapping(report_template_id=tid, **body.model_dump())
    session.add(mapping)
    session.commit()
    session.refresh(mapping)
    return _mapping_public(mapping)


@router.post("/{id}/templates/{tid}/mappings/update", response_model=SheetMappingPublic)
def update_mapping(session: SessionDep, id: uuid.UUID, tid: uuid.UUID, body: SheetMappingUpdate,
    _: User = Depends(require_permission_for_resource(RES, PermissionActionEnum.UPDATE, _mid_from_path))) -> Any:
    mapping = session.get(ReportSheetMapping, body.id)
    if not mapping or mapping.report_template_id != tid:
        raise HTTPException(404, "Mapping not found")
    data = body.model_dump(exclude_unset=True, exclude={"id"})
    mapping.sqlmodel_update(data)
    mapping.updated_at = datetime.now(UTC)
    session.add(mapping)
    session.commit()
    session.refresh(mapping)
    return _mapping_public(mapping)


@router.post("/{id}/templates/{tid}/mappings/batch-update", response_model=list[SheetMappingPublic])
def batch_update_mappings(session: SessionDep, id: uuid.UUID, tid: uuid.UUID, body: SheetMappingBatchUpdate,
    _: User = Depends(require_permission_for_resource(RES, PermissionActionEnum.UPDATE, _mid_from_path))) -> Any:
    if not body.mappings:
        raise HTTPException(400, "No mappings provided")
    updated = []
    now = datetime.now(UTC)
    for item in body.mappings:
        mapping = session.get(ReportSheetMapping, item.id)
        if not mapping or mapping.report_template_id != tid:
            raise HTTPException(404, f"Mapping {item.id} not found")
        data = item.model_dump(exclude_unset=True, exclude={"id"})
        mapping.sqlmodel_update(data)
        mapping.updated_at = now
        session.add(mapping)
        updated.append(mapping)
    session.commit()
    for m in updated:
        session.refresh(m)
    return [_mapping_public(m) for m in updated]


@router.post("/{id}/templates/{tid}/mappings/delete", response_model=Message)
def delete_mapping(session: SessionDep, id: uuid.UUID, tid: uuid.UUID, mapping_id: uuid.UUID,
    _: User = Depends(require_permission_for_resource(RES, PermissionActionEnum.UPDATE, _mid_from_path))) -> Any:
    mapping = session.get(ReportSheetMapping, mapping_id)
    if not mapping or mapping.report_template_id != tid:
        raise HTTPException(404, "Mapping not found")
    session.delete(mapping)
    session.commit()
    return Message(message="Mapping deleted")


# ========================= Generate + Executions =========================



def _parse_token(request: Request) -> dict:
    """Extract and decode JWT from Authorization header. Returns payload dict."""
    import jwt as pyjwt

    from app.core.security import ALGORITHM

    auth = request.headers.get("authorization", "")
    token = auth.replace("Bearer ", "").replace("bearer ", "").strip()
    if not token:
        raise HTTPException(401, "Authorization required")
    try:
        return pyjwt.decode(token, settings.SECRET_KEY, algorithms=[ALGORITHM], options={"verify_exp": True})
    except pyjwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except pyjwt.InvalidTokenError:
        raise HTTPException(403, "Invalid token")


def _get_client_module_ids(session: Session, payload: dict) -> list[uuid.UUID] | None:
    """Return list of module IDs the token has access to.

    Dashboard token → None (all modules).
    Gateway client token → list of assigned module IDs.
    """
    from app.core.security import TOKEN_TYPE_DASHBOARD, TOKEN_TYPE_GATEWAY
    from app.models_dbapi import AppClient

    token_type = payload.get("type")
    if token_type == TOKEN_TYPE_DASHBOARD:
        return None

    if token_type == TOKEN_TYPE_GATEWAY:
        client_id_str = payload.get("sub")
        if not client_id_str:
            raise HTTPException(403, "Invalid client token")
        client = session.exec(select(AppClient).where(AppClient.client_id == client_id_str)).first()
        if not client:
            raise HTTPException(403, "Client not found")
        rows = session.exec(
            select(ReportModuleClientLink.report_module_id).where(
                ReportModuleClientLink.app_client_id == client.id,
            )
        ).all()
        return list(rows)

    raise HTTPException(403, "Invalid token type")


def _verify_report_access(session: Session, module_id: uuid.UUID, token: str) -> None:
    """Verify token: dashboard user (any) OR gateway client (assigned to module)."""
    import jwt as pyjwt

    from app.core.security import ALGORITHM

    try:
        payload = pyjwt.decode(token, settings.SECRET_KEY, algorithms=[ALGORITHM], options={"verify_exp": True})
    except pyjwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except pyjwt.InvalidTokenError:
        raise HTTPException(403, "Invalid token")

    allowed = _get_client_module_ids(session, payload)
    if allowed is not None and module_id not in allowed:
        raise HTTPException(403, "Client not authorized for this report module")


@router.post("/{id}/templates/{tid}/generate", response_model=ReportGenerateOut)
def generate(session: SessionDep, request: Request, id: uuid.UUID, tid: uuid.UUID, body: ReportGenerateIn) -> Any:
    # Accept both dashboard and gateway client tokens
    auth = request.headers.get("authorization", "")
    token = auth.replace("Bearer ", "").replace("bearer ", "").strip()
    if not token:
        raise HTTPException(401, "Authorization required")
    _verify_report_access(session, id, token)

    mod = session.get(ReportModule, id)
    if not mod or not mod.is_active:
        raise HTTPException(400, "Module not found or inactive")
    tpl = session.get(ReportTemplate, tid)
    if not tpl or tpl.report_module_id != id or not tpl.is_active:
        raise HTTPException(400, "Template not found or inactive")
    mappings = list(session.exec(select(ReportSheetMapping).where(
        ReportSheetMapping.report_template_id == tid, ReportSheetMapping.is_active == True)).all())
    if not mappings:
        raise HTTPException(400, "No active mappings")

    exc = ReportExecution(report_template_id=tid, status=ExecutionStatusEnum.PENDING, parameters=body.parameters)
    session.add(exc)
    session.commit()
    session.refresh(exc)

    if body.is_async:
        from app.engines.excel.job_queue import enqueue_report_job
        enqueued = enqueue_report_job(exc.id, mod.id, tpl.id, body.parameters)
        if not enqueued:
            # Fallback: Redis unavailable, run inline
            _log.warning("Redis unavailable, running report sync instead of async")
            ExcelReportExecutor().execute(session, mod, tpl, mappings, exc, body.parameters)
            session.refresh(exc)
            return ReportGenerateOut(execution_id=exc.id, status=exc.status, output_url=exc.output_url, output_minio_path=exc.output_minio_path)
        return ReportGenerateOut(execution_id=exc.id, status=ExecutionStatusEnum.PENDING)

    ExcelReportExecutor().execute(session, mod, tpl, mappings, exc, body.parameters)
    session.refresh(exc)
    return ReportGenerateOut(execution_id=exc.id, status=exc.status, output_url=exc.output_url, output_minio_path=exc.output_minio_path)


@router.get("/{id}/templates/{tid}/executions", response_model=ReportExecutionListOut)
def list_template_executions(session: SessionDep, id: uuid.UUID, tid: uuid.UUID,
    page: int = 1, page_size: int = 20,
    _: User = Depends(require_permission_for_resource(RES, PermissionActionEnum.READ, _mid_from_path))) -> Any:
    total = session.exec(select(func.count()).select_from(ReportExecution).where(ReportExecution.report_template_id == tid)).one()
    offset = (page - 1) * page_size
    rows = session.exec(select(ReportExecution).where(ReportExecution.report_template_id == tid).order_by(ReportExecution.created_at.desc()).offset(offset).limit(page_size)).all()
    return ReportExecutionListOut(data=[_execution_public(e) for e in rows], total=total)


# ========================= Client-accessible endpoints =========================
# Accept both Dashboard JWT and Client JWT (gateway token).
# Client tokens are scoped to assigned modules only.
# Read + CRUD on templates/mappings within assigned modules.
# Module CRUD and user management remain Dashboard-only.


def _require_client_module_access(
    session: Session, payload: dict, module_id: uuid.UUID,
) -> None:
    """Raise 403 if client token is not assigned to this module."""
    allowed = _get_client_module_ids(session, payload)
    if allowed is not None and module_id not in allowed:
        raise HTTPException(403, "Not authorized for this module")


@router.post("/client/modules", response_model=ReportModuleListOut)
def client_list_modules(session: SessionDep, request: Request, body: ReportModuleListIn) -> Any:
    """List modules accessible by the current token (dashboard=all, client=assigned only)."""
    payload = _parse_token(request)
    allowed = _get_client_module_ids(session, payload)

    stmt = select(ReportModule).where(ReportModule.is_active == True)  # noqa: E712
    if allowed is not None:
        stmt = stmt.where(ReportModule.id.in_(allowed))
    if body.name__ilike:
        stmt = stmt.where(ReportModule.name.ilike(f"%{body.name__ilike}%"))

    count_stmt = select(func.count()).select_from(stmt.subquery())
    total = session.exec(count_stmt).one()

    offset = ((body.page or 1) - 1) * (body.page_size or 20)
    rows = session.exec(stmt.order_by(ReportModule.name).offset(offset).limit(body.page_size or 20)).all()
    return ReportModuleListOut(data=[_module_public(m) for m in rows], total=total)


@router.post("/client/templates", response_model=ReportTemplateListOut)
def client_list_templates(session: SessionDep, request: Request, body: ReportTemplateListIn) -> Any:
    """List templates accessible by the current token."""
    payload = _parse_token(request)
    allowed = _get_client_module_ids(session, payload)

    stmt = select(ReportTemplate).where(ReportTemplate.is_active == True)  # noqa: E712
    if allowed is not None:
        stmt = stmt.where(ReportTemplate.report_module_id.in_(allowed))
    if body.name__ilike:
        stmt = stmt.where(ReportTemplate.name.ilike(f"%{body.name__ilike}%"))
    if body.module_id:
        stmt = stmt.where(ReportTemplate.report_module_id == body.module_id)

    count_stmt = select(func.count()).select_from(stmt.subquery())
    total = session.exec(count_stmt).one()

    offset = ((body.page or 1) - 1) * (body.page_size or 20)
    rows = session.exec(stmt.order_by(ReportTemplate.name).offset(offset).limit(body.page_size or 20)).all()
    return ReportTemplateListOut(data=[_template_public(t) for t in rows], total=total)


@router.get("/client/template-detail", response_model=ReportTemplateDetail)
def client_get_template(session: SessionDep, request: Request, tid: uuid.UUID) -> Any:
    """Get template detail. Client must be assigned to the template's module."""
    payload = _parse_token(request)
    tpl = session.get(ReportTemplate, tid)
    if not tpl:
        raise HTTPException(404, "Template not found")
    allowed = _get_client_module_ids(session, payload)
    if allowed is not None and tpl.report_module_id not in allowed:
        raise HTTPException(403, "Not authorized for this template")
    return _template_detail(tpl)


@router.get("/client/execution-detail", response_model=ReportExecutionPublic)
def client_get_execution(session: SessionDep, request: Request, exec_id: uuid.UUID) -> Any:
    """Get execution status. Client must be assigned to the execution's module."""
    payload = _parse_token(request)
    exc = session.get(ReportExecution, exec_id)
    if not exc:
        raise HTTPException(404, "Execution not found")
    tpl = session.get(ReportTemplate, exc.report_template_id)
    if not tpl:
        raise HTTPException(404, "Execution not found")
    allowed = _get_client_module_ids(session, payload)
    if allowed is not None and tpl.report_module_id not in allowed:
        raise HTTPException(403, "Not authorized")
    return _execution_public(exc)


@router.get("/client/executions", response_model=ReportExecutionListOut)
def client_list_executions(
    session: SessionDep, request: Request,
    page: int = 1, page_size: int = 20,
    status: ExecutionStatusEnum | None = None,
    template_id: uuid.UUID | None = None,
    module_id: uuid.UUID | None = None,
) -> Any:
    """List executions accessible by the current token."""
    payload = _parse_token(request)
    allowed = _get_client_module_ids(session, payload)

    wheres = []
    if status:
        wheres.append(ReportExecution.status == status)
    if template_id:
        wheres.append(ReportExecution.report_template_id == template_id)
    if module_id:
        wheres.append(ReportTemplate.report_module_id == module_id)
    if allowed is not None:
        wheres.append(ReportTemplate.report_module_id.in_(allowed))

    count_stmt = select(func.count()).select_from(ReportExecution).join(
        ReportTemplate, ReportExecution.report_template_id == ReportTemplate.id,
    )
    for w in wheres:
        count_stmt = count_stmt.where(w)
    total = session.exec(count_stmt).one()

    stmt = select(ReportExecution).join(
        ReportTemplate, ReportExecution.report_template_id == ReportTemplate.id,
    )
    for w in wheres:
        stmt = stmt.where(w)
    offset = (page - 1) * page_size
    rows = session.exec(stmt.order_by(ReportExecution.created_at.desc()).offset(offset).limit(page_size)).all()
    return ReportExecutionListOut(data=[_execution_public(e) for e in rows], total=total)


# --- Client: Module detail ---


@router.get("/client/module-detail", response_model=ReportModuleDetail)
def client_get_module(session: SessionDep, request: Request, mid: uuid.UUID) -> Any:
    payload = _parse_token(request)
    _require_client_module_access(session, payload, mid)
    mod = session.get(ReportModule, mid)
    if not mod:
        raise HTTPException(404, "Module not found")
    return _module_detail(mod)


# --- Client: Template CRUD ---


@router.post("/client/modules/{mid}/templates/create", response_model=ReportTemplateDetail)
def client_create_template(
    session: SessionDep, request: Request, mid: uuid.UUID, body: ReportTemplateCreate,
) -> Any:
    payload = _parse_token(request)
    _require_client_module_access(session, payload, mid)
    mod = session.get(ReportModule, mid)
    if not mod:
        raise HTTPException(404, "Module not found")
    tpl_data = body.model_dump(exclude={"sheet_mappings"})
    tpl = ReportTemplate(report_module_id=mid, **tpl_data)
    session.add(tpl)
    session.flush()
    if body.sheet_mappings:
        for m in body.sheet_mappings:
            session.add(ReportSheetMapping(report_template_id=tpl.id, **m.model_dump()))
    session.commit()
    session.refresh(tpl)
    return _template_detail(tpl)


@router.post("/client/modules/{mid}/templates/update", response_model=ReportTemplatePublic)
def client_update_template(
    session: SessionDep, request: Request, mid: uuid.UUID, body: ReportTemplateUpdate,
) -> Any:
    payload = _parse_token(request)
    _require_client_module_access(session, payload, mid)
    tpl = session.get(ReportTemplate, body.id)
    if not tpl or tpl.report_module_id != mid:
        raise HTTPException(404, "Template not found")
    data = body.model_dump(exclude_unset=True, exclude={"id"})
    tpl.sqlmodel_update(data)
    tpl.updated_at = datetime.now(UTC)
    session.add(tpl)
    session.commit()
    session.refresh(tpl)
    return _template_public(tpl)


@router.post("/client/modules/{mid}/templates/delete", response_model=Message)
def client_delete_template(
    session: SessionDep, request: Request, mid: uuid.UUID, tid: uuid.UUID,
) -> Any:
    payload = _parse_token(request)
    _require_client_module_access(session, payload, mid)
    tpl = session.get(ReportTemplate, tid)
    if not tpl or tpl.report_module_id != mid:
        raise HTTPException(404, "Template not found")
    session.delete(tpl)
    session.commit()
    return Message(message="Template deleted")


# --- Client: Mapping CRUD ---


@router.post("/client/modules/{mid}/templates/{tid}/mappings/create", response_model=SheetMappingPublic)
def client_create_mapping(
    session: SessionDep, request: Request, mid: uuid.UUID, tid: uuid.UUID,
    body: SheetMappingCreate,
) -> Any:
    payload = _parse_token(request)
    _require_client_module_access(session, payload, mid)
    tpl = session.get(ReportTemplate, tid)
    if not tpl or tpl.report_module_id != mid:
        raise HTTPException(404, "Template not found")
    mapping = ReportSheetMapping(report_template_id=tid, **body.model_dump())
    session.add(mapping)
    session.commit()
    session.refresh(mapping)
    return _mapping_public(mapping)


@router.post("/client/modules/{mid}/templates/{tid}/mappings/update", response_model=SheetMappingPublic)
def client_update_mapping(
    session: SessionDep, request: Request, mid: uuid.UUID, tid: uuid.UUID,
    body: SheetMappingUpdate,
) -> Any:
    payload = _parse_token(request)
    _require_client_module_access(session, payload, mid)
    mapping = session.get(ReportSheetMapping, body.id)
    if not mapping or mapping.report_template_id != tid:
        raise HTTPException(404, "Mapping not found")
    tpl = session.get(ReportTemplate, tid)
    if not tpl or tpl.report_module_id != mid:
        raise HTTPException(404, "Template not found")
    data = body.model_dump(exclude_unset=True, exclude={"id"})
    mapping.sqlmodel_update(data)
    mapping.updated_at = datetime.now(UTC)
    session.add(mapping)
    session.commit()
    session.refresh(mapping)
    return _mapping_public(mapping)


@router.post("/client/modules/{mid}/templates/{tid}/mappings/batch-update", response_model=list[SheetMappingPublic])
def client_batch_update_mappings(
    session: SessionDep, request: Request, mid: uuid.UUID, tid: uuid.UUID,
    body: SheetMappingBatchUpdate,
) -> Any:
    payload = _parse_token(request)
    _require_client_module_access(session, payload, mid)
    tpl = session.get(ReportTemplate, tid)
    if not tpl or tpl.report_module_id != mid:
        raise HTTPException(404, "Template not found")
    if not body.mappings:
        raise HTTPException(400, "No mappings provided")
    updated = []
    now = datetime.now(UTC)
    for item in body.mappings:
        mapping = session.get(ReportSheetMapping, item.id)
        if not mapping or mapping.report_template_id != tid:
            raise HTTPException(404, f"Mapping {item.id} not found")
        data = item.model_dump(exclude_unset=True, exclude={"id"})
        mapping.sqlmodel_update(data)
        mapping.updated_at = now
        session.add(mapping)
        updated.append(mapping)
    session.commit()
    for m in updated:
        session.refresh(m)
    return [_mapping_public(m) for m in updated]


@router.post("/client/modules/{mid}/templates/{tid}/mappings/delete", response_model=Message)
def client_delete_mapping(
    session: SessionDep, request: Request, mid: uuid.UUID, tid: uuid.UUID,
    mapping_id: uuid.UUID,
) -> Any:
    payload = _parse_token(request)
    _require_client_module_access(session, payload, mid)
    mapping = session.get(ReportSheetMapping, mapping_id)
    if not mapping or mapping.report_template_id != tid:
        raise HTTPException(404, "Mapping not found")
    tpl = session.get(ReportTemplate, tid)
    if not tpl or tpl.report_module_id != mid:
        raise HTTPException(404, "Template not found")
    session.delete(mapping)
    session.commit()
    return Message(message="Mapping deleted")


# --- Client: Template client access ---


@router.get("/client/modules/{mid}/templates/{tid}/clients", response_model=ClientIdsOut)
def client_get_template_clients(
    session: SessionDep, request: Request, mid: uuid.UUID, tid: uuid.UUID,
) -> Any:
    payload = _parse_token(request)
    _require_client_module_access(session, payload, mid)
    tpl = session.get(ReportTemplate, tid)
    if not tpl or tpl.report_module_id != mid:
        raise HTTPException(404, "Template not found")
    return ClientIdsOut(client_ids=[lnk.app_client_id for lnk in (tpl.client_links or [])])


@router.post("/client/modules/{mid}/templates/{tid}/clients", response_model=ClientIdsOut)
def client_set_template_clients(
    session: SessionDep, request: Request, mid: uuid.UUID, tid: uuid.UUID,
    body: ClientIdsUpdate,
) -> Any:
    payload = _parse_token(request)
    _require_client_module_access(session, payload, mid)
    tpl = session.get(ReportTemplate, tid)
    if not tpl or tpl.report_module_id != mid:
        raise HTTPException(404, "Template not found")
    for lnk in session.exec(select(ReportTemplateClientLink).where(ReportTemplateClientLink.report_template_id == tid)).all():
        session.delete(lnk)
    for cid in body.client_ids:
        session.add(ReportTemplateClientLink(report_template_id=tid, app_client_id=cid))
    session.commit()
    session.refresh(tpl)
    return ClientIdsOut(client_ids=[lnk.app_client_id for lnk in (tpl.client_links or [])])


# --- Client: Template executions ---


@router.get("/client/modules/{mid}/templates/{tid}/executions", response_model=ReportExecutionListOut)
def client_list_template_executions(
    session: SessionDep, request: Request, mid: uuid.UUID, tid: uuid.UUID,
    page: int = 1, page_size: int = 20,
) -> Any:
    payload = _parse_token(request)
    _require_client_module_access(session, payload, mid)
    tpl = session.get(ReportTemplate, tid)
    if not tpl or tpl.report_module_id != mid:
        raise HTTPException(404, "Template not found")
    total = session.exec(select(func.count()).select_from(ReportExecution).where(ReportExecution.report_template_id == tid)).one()
    offset = (page - 1) * page_size
    rows = session.exec(select(ReportExecution).where(ReportExecution.report_template_id == tid).order_by(ReportExecution.created_at.desc()).offset(offset).limit(page_size)).all()
    return ReportExecutionListOut(data=[_execution_public(e) for e in rows], total=total)


# --- Client: MinIO helpers ---


@router.get("/client/buckets/{datasource_id}", response_model=list[str])
def client_list_buckets(session: SessionDep, request: Request, datasource_id: uuid.UUID) -> Any:
    _parse_token(request)
    from app.engines.excel.minio_client import get_minio_client
    from app.models_dbapi import DataSource

    ds = session.get(DataSource, datasource_id)
    if not ds:
        raise HTTPException(404, "Datasource not found")
    try:
        client = get_minio_client(ds)
        return sorted([b.name for b in client.list_buckets()])
    except Exception as e:
        raise HTTPException(400, f"Failed to list buckets: {e}")


@router.get("/client/files/{datasource_id}/{bucket}", response_model=list[str])
def client_list_files(
    session: SessionDep, request: Request, datasource_id: uuid.UUID,
    bucket: str, prefix: str = "", suffix: str = ".xlsx",
) -> Any:
    _parse_token(request)
    from app.engines.excel.minio_client import get_minio_client
    from app.models_dbapi import DataSource

    _validate_path(bucket)
    if prefix:
        _validate_path(prefix)
    ds = session.get(DataSource, datasource_id)
    if not ds:
        raise HTTPException(404, "Datasource not found")
    try:
        client = get_minio_client(ds)
        objects = client.list_objects(bucket, prefix=prefix or None, recursive=True)
        return sorted([o.object_name for o in objects if not o.is_dir and o.object_name.endswith(suffix)])
    except Exception as e:
        raise HTTPException(400, f"Failed to list files: {e}")


@router.get("/client/sheets/{datasource_id}/{bucket}/{file_path:path}", response_model=list[str])
def client_list_sheets(
    session: SessionDep, request: Request, datasource_id: uuid.UUID,
    bucket: str, file_path: str,
) -> Any:
    _parse_token(request)
    import os
    import tempfile

    from openpyxl import load_workbook

    from app.engines.excel.minio_client import download_file, get_minio_client
    from app.models_dbapi import DataSource

    _validate_path(bucket)
    _validate_path(file_path)
    ds = session.get(DataSource, datasource_id)
    if not ds:
        raise HTTPException(404, "Datasource not found")
    tmp_path = None
    try:
        client = get_minio_client(ds)
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp_path = tmp.name
        tmp.close()
        download_file(client, bucket, file_path, tmp_path)
        wb = load_workbook(tmp_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        return sheets
    except HTTPException:
        raise
    except Exception as e:
        _log.warning("Failed to read sheets: %s", e)
        raise HTTPException(400, "Unable to read sheets")
    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)
