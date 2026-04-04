"""Seed report module examples with templates and mappings.

Controlled by SEED_EXAMPLE_DATA. Idempotent: skips if report module already exists.
Creates a MinIO datasource, report module, buckets, template files, and sample templates.

Templates created:
  1. product-list          - Simple rows export from template
  2. order-summary         - Recalc + extract sheet (formulas)
  3. multi-query-sheet     - Multiple mappings on the same sheet (single + rows)
  4. blank-users           - Blank workbook, single sheet
  5. multi-sheet-blank     - Blank workbook, multiple sheets
  6. parameterized-report  - Jinja2 SQL with dynamic parameters

Optionally creates a StarRocks variant module if a StarRocks datasource exists.
"""

from __future__ import annotations

import io
import logging

from openpyxl import Workbook
from sqlmodel import Session, select

from app.core.config import settings
from app.core.security import encrypt_value
from app.core.permission_resources import ensure_resource_permissions
from app.models_dbapi import AppClient, DataSource, ProductTypeEnum
from app.models_permission import PermissionActionEnum, ResourceTypeEnum
from app.models_report import (
    ReportModule,
    ReportModuleClientLink,
    ReportSheetMapping,
    ReportTemplate,
    SheetWriteModeEnum,
)

logger = logging.getLogger(__name__)

REPORT_MODULE_NAME = "Examples (Reports)"
REPORT_MODULE_SR_NAME = "Examples (Reports - StarRocks)"
MINIO_DS_NAME = "Examples MinIO"

_RM_ACTIONS = (
    PermissionActionEnum.READ,
    PermissionActionEnum.CREATE,
    PermissionActionEnum.UPDATE,
    PermissionActionEnum.DELETE,
    PermissionActionEnum.EXECUTE,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _already_seeded(session: Session, name: str) -> bool:
    return (
        session.exec(
            select(ReportModule).where(ReportModule.name == name)
        ).first()
        is not None
    )


def _get_or_create_minio_ds(session: Session) -> DataSource:
    """Get or create MinIO datasource pointing to the docker minio service."""
    existing = session.exec(
        select(DataSource).where(DataSource.name == MINIO_DS_NAME)
    ).first()
    if existing:
        return existing

    ds = DataSource(
        name=MINIO_DS_NAME,
        product_type=ProductTypeEnum.MINIO,
        host="minio",
        port=9000,
        database="",
        username="minioadmin",
        password=encrypt_value("minioadmin"),
        description="MinIO storage for report examples (auto-seeded)",
        use_ssl=False,
    )
    session.add(ds)
    session.flush()
    ensure_resource_permissions(
        session, ResourceTypeEnum.DATASOURCE, ds.id, _RM_ACTIONS
    )
    logger.info("Created MinIO datasource: %s", ds.name)
    return ds


def _get_sql_datasource(session: Session) -> DataSource | None:
    """Find an existing PostgreSQL datasource."""
    return session.exec(
        select(DataSource).where(DataSource.product_type == ProductTypeEnum.POSTGRES)
    ).first()


def _get_starrocks_datasource(session: Session) -> DataSource | None:
    """Find an existing StarRocks (MySQL-protocol) datasource, if any."""
    # StarRocks datasources use the MySQL product type
    ds = session.exec(
        select(DataSource).where(DataSource.product_type == ProductTypeEnum.MYSQL)
    ).first()
    if ds and ds.description and "starrocks" in ds.description.lower():
        return ds
    # Also check via STARROCKS_HOST env var if set
    sr_host = getattr(settings, "STARROCKS_HOST", None)
    if sr_host:
        return session.exec(
            select(DataSource).where(
                DataSource.product_type == ProductTypeEnum.MYSQL,
                DataSource.host == sr_host,
            )
        ).first()
    return ds  # Return MySQL ds if found (may or may not be StarRocks)


def _get_client_by_client_id(session: Session, client_id: str) -> AppClient | None:
    """Look up an AppClient by its client_id string."""
    return session.exec(
        select(AppClient).where(AppClient.client_id == client_id)
    ).first()


# ---------------------------------------------------------------------------
# MinIO Setup
# ---------------------------------------------------------------------------


def _setup_minio_buckets_and_templates(ds: DataSource) -> None:
    """Create buckets and upload sample Excel templates to MinIO."""
    from app.engines.excel.minio_client import get_minio_client

    client = get_minio_client(ds)

    # Create buckets
    for bucket in ["report-templates", "report-output"]:
        if not client.bucket_exists(bucket):
            client.make_bucket(bucket)
            logger.info("Created MinIO bucket: %s", bucket)

    # ---------------------------------------------------------------
    # Template 1: products/product-list.xlsx
    # Simple template with header row, Data sheet
    # ---------------------------------------------------------------
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "Data"
    ws1["A1"] = "Report: Product List"
    ws1["A2"] = "Generated by pyDBAPI Report Engine"
    ws1["A4"] = "id"
    ws1["B4"] = "name"
    ws1["C4"] = "price"
    ws1["D4"] = "category"
    buf1 = io.BytesIO()
    wb1.save(buf1)
    buf1.seek(0)
    client.put_object(
        "report-templates", "products/product-list.xlsx", buf1, len(buf1.getvalue())
    )
    wb1.close()
    logger.info("Uploaded template: products/product-list.xlsx")

    # ---------------------------------------------------------------
    # Template 2: orders/order-summary.xlsx
    # RawData sheet + Summary sheet with formulas (COUNTA, SUM, AVERAGE, NOW)
    # ---------------------------------------------------------------
    wb2 = Workbook()
    ws_data = wb2.active
    ws_data.title = "RawData"
    ws_data["A1"] = "order_id"
    ws_data["B1"] = "status"
    ws_data["C1"] = "total"

    ws_summary = wb2.create_sheet("Summary")
    ws_summary["A1"] = "Order Summary Report"
    ws_summary["A3"] = "Total Orders:"
    ws_summary["B3"] = "=COUNTA(RawData!A2:A1000)"
    ws_summary["A4"] = "Total Revenue:"
    ws_summary["B4"] = "=SUM(RawData!C2:C1000)"
    ws_summary["A5"] = "Average Order:"
    ws_summary["B5"] = "=IF(B3>0, B4/B3, 0)"
    ws_summary["A7"] = "Generated at:"
    ws_summary["B7"] = "=NOW()"

    buf2 = io.BytesIO()
    wb2.save(buf2)
    buf2.seek(0)
    client.put_object(
        "report-templates", "orders/order-summary.xlsx", buf2, len(buf2.getvalue())
    )
    wb2.close()
    logger.info("Uploaded template: orders/order-summary.xlsx")

    # ---------------------------------------------------------------
    # Template 3: combined/multi-query-sheet.xlsx
    # Single sheet "Report" with multiple data areas:
    #   A1: header title
    #   E1: placeholder for product count (single)
    #   E2: placeholder for order total (single)
    #   A4: column headers for product listing
    #   A5+: rows start
    # ---------------------------------------------------------------
    wb3 = Workbook()
    ws_report = wb3.active
    ws_report.title = "Report"
    ws_report["A1"] = "Combined Report"
    ws_report["D1"] = "Products:"
    ws_report["E1"] = ""  # placeholder: will be filled by single mapping
    ws_report["D2"] = "Revenue:"
    ws_report["E2"] = ""  # placeholder: will be filled by single mapping
    ws_report["A4"] = "id"
    ws_report["B4"] = "name"
    ws_report["C4"] = "price"

    buf3 = io.BytesIO()
    wb3.save(buf3)
    buf3.seek(0)
    client.put_object(
        "report-templates", "combined/multi-query-sheet.xlsx", buf3, len(buf3.getvalue())
    )
    wb3.close()
    logger.info("Uploaded template: combined/multi-query-sheet.xlsx")


# ---------------------------------------------------------------------------
# Template + Mapping Creation
# ---------------------------------------------------------------------------


def _create_templates_postgres(session: Session, mod: ReportModule) -> None:
    """Create all 6 template definitions for a PostgreSQL-backed module."""

    # --- Template 1: product-list (simple rows export) ---
    tpl1 = ReportTemplate(
        report_module_id=mod.id,
        name="product-list",
        description="Simple product catalog export",
        template_bucket="report-templates",
        template_path="products/product-list.xlsx",
        output_bucket="report-output",
        output_prefix="products/",
        recalc_enabled=False,
    )
    session.add(tpl1)
    session.flush()
    session.add(
        ReportSheetMapping(
            report_template_id=tpl1.id,
            sort_order=1,
            sheet_name="Data",
            start_cell="A5",
            write_mode=SheetWriteModeEnum.ROWS,
            write_headers=True,
            sql_content="SELECT id, name, price, category FROM products ORDER BY id",
            description="All products",
        )
    )
    logger.info("Created template: product-list (1 mapping)")

    # --- Template 2: order-summary (recalc + extract sheet) ---
    tpl2 = ReportTemplate(
        report_module_id=mod.id,
        name="order-summary",
        description="Order summary with formula recalculation",
        template_bucket="report-templates",
        template_path="orders/order-summary.xlsx",
        output_bucket="report-output",
        output_prefix="orders/",
        recalc_enabled=True,
        output_sheet="Summary",
    )
    session.add(tpl2)
    session.flush()
    session.add(
        ReportSheetMapping(
            report_template_id=tpl2.id,
            sort_order=1,
            sheet_name="RawData",
            start_cell="A2",
            write_mode=SheetWriteModeEnum.ROWS,
            write_headers=False,
            sql_content="SELECT id, status, total FROM orders ORDER BY id",
            description="Order data for formula sheet",
        )
    )
    logger.info(
        "Created template: order-summary (1 mapping, recalc=true, output_sheet=Summary)"
    )

    # --- Template 3: multi-query-sheet (multiple mappings on same sheet) ---
    tpl3 = ReportTemplate(
        report_module_id=mod.id,
        name="multi-query-sheet",
        description="Multiple data areas on a single sheet (single + rows)",
        template_bucket="report-templates",
        template_path="combined/multi-query-sheet.xlsx",
        output_bucket="report-output",
        output_prefix="combined/",
        recalc_enabled=False,
    )
    session.add(tpl3)
    session.flush()
    # Mapping 3a: product count at E1 (single)
    session.add(
        ReportSheetMapping(
            report_template_id=tpl3.id,
            sort_order=1,
            sheet_name="Report",
            start_cell="E1",
            write_mode=SheetWriteModeEnum.SINGLE,
            write_headers=False,
            sql_content="SELECT COUNT(*) FROM products",
            description="Total product count",
        )
    )
    # Mapping 3b: order revenue at E2 (single)
    session.add(
        ReportSheetMapping(
            report_template_id=tpl3.id,
            sort_order=2,
            sheet_name="Report",
            start_cell="E2",
            write_mode=SheetWriteModeEnum.SINGLE,
            write_headers=False,
            sql_content="SELECT SUM(total) FROM orders",
            description="Total order revenue",
        )
    )
    # Mapping 3c: product listing at A5 (rows)
    session.add(
        ReportSheetMapping(
            report_template_id=tpl3.id,
            sort_order=3,
            sheet_name="Report",
            start_cell="A5",
            write_mode=SheetWriteModeEnum.ROWS,
            write_headers=True,
            sql_content="SELECT id, name, price FROM products ORDER BY id",
            description="Product listing",
        )
    )
    logger.info("Created template: multi-query-sheet (3 mappings on same sheet)")

    # --- Template 4: blank-users (blank workbook, no template file) ---
    tpl4 = ReportTemplate(
        report_module_id=mod.id,
        name="blank-users",
        description="User export from blank workbook (no template file)",
        template_bucket="",
        template_path="",
        output_bucket="report-output",
        output_prefix="users/",
        recalc_enabled=False,
    )
    session.add(tpl4)
    session.flush()
    session.add(
        ReportSheetMapping(
            report_template_id=tpl4.id,
            sort_order=1,
            sheet_name="Users",
            start_cell="A1",
            write_mode=SheetWriteModeEnum.ROWS,
            write_headers=True,
            sql_content="SELECT id, username, email, is_active FROM sample_users ORDER BY id",
            description="All users",
        )
    )
    logger.info("Created template: blank-users (1 mapping, blank template)")

    # --- Template 5: multi-sheet-blank (multi-sheet from blank workbook) ---
    tpl5 = ReportTemplate(
        report_module_id=mod.id,
        name="multi-sheet-blank",
        description="Multi-sheet report generated from blank workbook",
        template_bucket="",
        template_path="",
        output_bucket="report-output",
        output_prefix="combined/",
        recalc_enabled=False,
    )
    session.add(tpl5)
    session.flush()
    session.add(
        ReportSheetMapping(
            report_template_id=tpl5.id,
            sort_order=1,
            sheet_name="Users",
            start_cell="A1",
            write_mode=SheetWriteModeEnum.ROWS,
            write_headers=True,
            sql_content="SELECT id, username, email FROM sample_users ORDER BY id",
            description="User list",
        )
    )
    session.add(
        ReportSheetMapping(
            report_template_id=tpl5.id,
            sort_order=2,
            sheet_name="Metrics",
            start_cell="A1",
            write_mode=SheetWriteModeEnum.ROWS,
            write_headers=True,
            sql_content="SELECT id, status, duration_ms, amount FROM metrics ORDER BY id",
            description="Metrics data",
        )
    )
    logger.info("Created template: multi-sheet-blank (2 mappings, blank template)")

    # --- Template 6: parameterized-report (Jinja2 SQL) ---
    tpl6 = ReportTemplate(
        report_module_id=mod.id,
        name="parameterized-report",
        description="Dynamic filtering with Jinja2 SQL parameters",
        template_bucket="",
        template_path="",
        output_bucket="report-output",
        output_prefix="parameterized/",
        recalc_enabled=False,
    )
    session.add(tpl6)
    session.flush()
    session.add(
        ReportSheetMapping(
            report_template_id=tpl6.id,
            sort_order=1,
            sheet_name="Data",
            start_cell="A1",
            write_mode=SheetWriteModeEnum.ROWS,
            write_headers=True,
            sql_content=(
                "SELECT id, name, price, category FROM products "
                "{% where %}"
                "{% if category %}AND category = {{ category | sql_string }}{% endif %}"
                "{% if min_price %}AND price >= {{ min_price | sql_float }}{% endif %}"
                "{% endwhere %} "
                "ORDER BY id"
            ),
            description="Products with optional category and min_price filters",
        )
    )
    logger.info("Created template: parameterized-report (1 mapping, Jinja2 SQL)")


def _create_templates_starrocks(session: Session, mod: ReportModule) -> None:
    """Create template definitions adapted for StarRocks SQL dialect.

    Key differences from PostgreSQL:
    - No ILIKE: use LOWER(col) LIKE LOWER(val)
    - Jinja2 {% where %} tag works the same, but ILIKE filters need adaptation
    """

    # --- Template 1: product-list ---
    tpl1 = ReportTemplate(
        report_module_id=mod.id,
        name="product-list",
        description="Simple product catalog export (StarRocks)",
        template_bucket="report-templates",
        template_path="products/product-list.xlsx",
        output_bucket="report-output",
        output_prefix="sr-products/",
        recalc_enabled=False,
    )
    session.add(tpl1)
    session.flush()
    session.add(
        ReportSheetMapping(
            report_template_id=tpl1.id,
            sort_order=1,
            sheet_name="Data",
            start_cell="A5",
            write_mode=SheetWriteModeEnum.ROWS,
            write_headers=True,
            sql_content="SELECT id, name, price, category FROM products ORDER BY id",
            description="All products (StarRocks)",
        )
    )
    logger.info("Created StarRocks template: product-list")

    # --- Template 2: order-summary ---
    tpl2 = ReportTemplate(
        report_module_id=mod.id,
        name="order-summary",
        description="Order summary with recalculation (StarRocks)",
        template_bucket="report-templates",
        template_path="orders/order-summary.xlsx",
        output_bucket="report-output",
        output_prefix="sr-orders/",
        recalc_enabled=True,
        output_sheet="Summary",
    )
    session.add(tpl2)
    session.flush()
    session.add(
        ReportSheetMapping(
            report_template_id=tpl2.id,
            sort_order=1,
            sheet_name="RawData",
            start_cell="A2",
            write_mode=SheetWriteModeEnum.ROWS,
            write_headers=False,
            sql_content="SELECT id, status, total FROM orders ORDER BY id",
            description="Order data for formula sheet (StarRocks)",
        )
    )
    logger.info("Created StarRocks template: order-summary")

    # --- Template 3: multi-query-sheet ---
    tpl3 = ReportTemplate(
        report_module_id=mod.id,
        name="multi-query-sheet",
        description="Multiple data areas on a single sheet (StarRocks)",
        template_bucket="report-templates",
        template_path="combined/multi-query-sheet.xlsx",
        output_bucket="report-output",
        output_prefix="sr-combined/",
        recalc_enabled=False,
    )
    session.add(tpl3)
    session.flush()
    session.add(
        ReportSheetMapping(
            report_template_id=tpl3.id,
            sort_order=1,
            sheet_name="Report",
            start_cell="E1",
            write_mode=SheetWriteModeEnum.SINGLE,
            write_headers=False,
            sql_content="SELECT COUNT(*) FROM products",
            description="Total product count (StarRocks)",
        )
    )
    session.add(
        ReportSheetMapping(
            report_template_id=tpl3.id,
            sort_order=2,
            sheet_name="Report",
            start_cell="E2",
            write_mode=SheetWriteModeEnum.SINGLE,
            write_headers=False,
            sql_content="SELECT SUM(total) FROM orders",
            description="Total order revenue (StarRocks)",
        )
    )
    session.add(
        ReportSheetMapping(
            report_template_id=tpl3.id,
            sort_order=3,
            sheet_name="Report",
            start_cell="A5",
            write_mode=SheetWriteModeEnum.ROWS,
            write_headers=True,
            sql_content="SELECT id, name, price FROM products ORDER BY id",
            description="Product listing (StarRocks)",
        )
    )
    logger.info("Created StarRocks template: multi-query-sheet")

    # --- Template 4: blank-users ---
    tpl4 = ReportTemplate(
        report_module_id=mod.id,
        name="blank-users",
        description="User export from blank workbook (StarRocks)",
        template_bucket="",
        template_path="",
        output_bucket="report-output",
        output_prefix="sr-users/",
        recalc_enabled=False,
    )
    session.add(tpl4)
    session.flush()
    session.add(
        ReportSheetMapping(
            report_template_id=tpl4.id,
            sort_order=1,
            sheet_name="Users",
            start_cell="A1",
            write_mode=SheetWriteModeEnum.ROWS,
            write_headers=True,
            sql_content="SELECT id, username, email, is_active FROM sample_users ORDER BY id",
            description="All users (StarRocks)",
        )
    )
    logger.info("Created StarRocks template: blank-users")

    # --- Template 5: multi-sheet-blank ---
    tpl5 = ReportTemplate(
        report_module_id=mod.id,
        name="multi-sheet-blank",
        description="Multi-sheet blank workbook (StarRocks)",
        template_bucket="",
        template_path="",
        output_bucket="report-output",
        output_prefix="sr-combined/",
        recalc_enabled=False,
    )
    session.add(tpl5)
    session.flush()
    session.add(
        ReportSheetMapping(
            report_template_id=tpl5.id,
            sort_order=1,
            sheet_name="Users",
            start_cell="A1",
            write_mode=SheetWriteModeEnum.ROWS,
            write_headers=True,
            sql_content="SELECT id, username, email FROM sample_users ORDER BY id",
            description="User list (StarRocks)",
        )
    )
    session.add(
        ReportSheetMapping(
            report_template_id=tpl5.id,
            sort_order=2,
            sheet_name="Metrics",
            start_cell="A1",
            write_mode=SheetWriteModeEnum.ROWS,
            write_headers=True,
            sql_content="SELECT id, status, duration_ms, amount FROM metrics ORDER BY id",
            description="Metrics data (StarRocks)",
        )
    )
    logger.info("Created StarRocks template: multi-sheet-blank")

    # --- Template 6: parameterized-report (StarRocks: LOWER+LIKE instead of ILIKE) ---
    tpl6 = ReportTemplate(
        report_module_id=mod.id,
        name="parameterized-report",
        description="Dynamic filtering with Jinja2 SQL (StarRocks)",
        template_bucket="",
        template_path="",
        output_bucket="report-output",
        output_prefix="sr-parameterized/",
        recalc_enabled=False,
    )
    session.add(tpl6)
    session.flush()
    session.add(
        ReportSheetMapping(
            report_template_id=tpl6.id,
            sort_order=1,
            sheet_name="Data",
            start_cell="A1",
            write_mode=SheetWriteModeEnum.ROWS,
            write_headers=True,
            sql_content=(
                "SELECT id, name, price, category FROM products "
                "{% where %}"
                "{% if category %}AND LOWER(category) = LOWER({{ category | sql_string }}){% endif %}"
                "{% if min_price %}AND price >= {{ min_price | sql_float }}{% endif %}"
                "{% endwhere %} "
                "ORDER BY id"
            ),
            description="Products with optional filters (StarRocks LOWER+LIKE)",
        )
    )
    logger.info("Created StarRocks template: parameterized-report")


# ---------------------------------------------------------------------------
# Module Creation
# ---------------------------------------------------------------------------


def _create_module(
    session: Session,
    name: str,
    description: str,
    minio_ds: DataSource,
    sql_ds: DataSource,
) -> ReportModule:
    """Create a ReportModule with permissions."""
    mod = ReportModule(
        name=name,
        description=description,
        minio_datasource_id=minio_ds.id,
        sql_datasource_id=sql_ds.id,
        default_template_bucket="report-templates",
        default_output_bucket="report-output",
    )
    session.add(mod)
    session.flush()
    ensure_resource_permissions(
        session, ResourceTypeEnum.REPORT_MODULE, mod.id, _RM_ACTIONS
    )
    logger.info("Created report module: %s", mod.name)
    return mod


def _assign_client_to_module(
    session: Session, mod: ReportModule, client_id_str: str
) -> None:
    """Assign an AppClient to a ReportModule via the link table."""
    client = _get_client_by_client_id(session, client_id_str)
    if not client:
        logger.warning(
            "Client '%s' not found — skipping module client assignment", client_id_str
        )
        return
    # Check if already linked
    existing = session.exec(
        select(ReportModuleClientLink).where(
            ReportModuleClientLink.report_module_id == mod.id,
            ReportModuleClientLink.app_client_id == client.id,
        )
    ).first()
    if existing:
        return
    session.add(
        ReportModuleClientLink(
            report_module_id=mod.id,
            app_client_id=client.id,
        )
    )
    logger.info("Assigned client '%s' to module '%s'", client_id_str, mod.name)


# ---------------------------------------------------------------------------
# Main Entry Point
# ---------------------------------------------------------------------------


def seed_report_examples(session: Session) -> None:
    """Seed report module, templates, and mappings."""
    if _already_seeded(session, REPORT_MODULE_NAME):
        logger.info("Report examples already seeded — skipping")
        return

    logger.info("Seeding report example data...")

    # 1. MinIO datasource
    minio_ds = _get_or_create_minio_ds(session)

    # 2. SQL datasource (reuse existing PostgreSQL)
    sql_ds = _get_sql_datasource(session)
    if not sql_ds:
        logger.warning("No PostgreSQL datasource found — skipping report seed")
        return

    # 3. Setup MinIO buckets and upload templates
    try:
        _setup_minio_buckets_and_templates(minio_ds)
    except Exception as e:
        logger.warning("MinIO setup failed (minio may not be running): %s", e)
        logger.info("Skipping report seed — start MinIO and re-seed")
        return

    # 4. Create PostgreSQL report module
    mod = _create_module(
        session,
        REPORT_MODULE_NAME,
        "Auto-seeded report examples with sample templates",
        minio_ds,
        sql_ds,
    )

    # 5. Create all 6 templates
    _create_templates_postgres(session, mod)

    # 6. Assign the "mobile-app" client to this module
    _assign_client_to_module(session, mod, "mobile-app")

    # 7. Optionally create StarRocks module
    sr_ds = _get_starrocks_datasource(session)
    if sr_ds and sr_ds.id != sql_ds.id:
        if not _already_seeded(session, REPORT_MODULE_SR_NAME):
            logger.info("StarRocks datasource found — creating StarRocks report module")
            sr_mod = _create_module(
                session,
                REPORT_MODULE_SR_NAME,
                "Auto-seeded report examples using StarRocks SQL datasource",
                minio_ds,
                sr_ds,
            )
            _create_templates_starrocks(session, sr_mod)
            _assign_client_to_module(session, sr_mod, "mobile-app")
        else:
            logger.info("StarRocks report module already exists — skipping")

    session.commit()
    logger.info("Report example data seeded successfully")
