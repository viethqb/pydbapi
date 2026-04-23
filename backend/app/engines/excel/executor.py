"""Excel report executor: orchestrates the full generate flow."""
import gc
import logging
import os
import shutil
from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlmodel import Session

from app.core.config import settings
from app.engines.excel import excel_writer
from app.engines.excel.excel_extractor import extract_sheet
from app.engines.excel.excel_writer import write_rows, write_single
from app.engines.excel.minio_client import download_file, get_minio_client, upload_file
from app.engines.excel.recalc import recalc_workbook
from app.engines.sql import SQLTemplateEngine, execute_sql
from app.models_dbapi import DataSource
from app.models_report import (
    ExecutionStatusEnum,
    ReportExecution,
    ReportModule,
    ReportSheetMapping,
    ReportTemplate,
    SheetWriteModeEnum,
)

_log = logging.getLogger(__name__)


def _update_progress(
    session: Session, execution: ReportExecution,
    processed_rows: int, pct: int | None,
) -> None:
    """Commit current progress to DB so async pollers can observe it."""
    execution.processed_rows = processed_rows
    execution.progress_pct = pct
    session.add(execution)
    try:
        session.commit()
    except Exception as e:
        _log.warning("Failed to commit progress update: %s", e)
        session.rollback()


def _deep_merge(base: dict | None, override: dict | None) -> dict | None:
    """Recursively merge two dicts. override wins per-key. None values in override are ignored."""
    if not base and not override:
        return None
    if not base:
        return dict(override or {})
    if not override:
        return dict(base)
    result: dict = {}
    for k in set(base) | set(override):
        bv, ov = base.get(k), override.get(k)
        if isinstance(bv, dict) and isinstance(ov, dict):
            result[k] = _deep_merge(bv, ov)
        elif ov is not None:
            result[k] = ov
        else:
            result[k] = bv
    return result


def _effective_start(
    sheet_name: str, start_cell: str, gap_rows: int,
    sheet_last_row: dict[str, int],
) -> tuple[str, bool]:
    """Resolve effective start_cell given prior writes on same sheet.

    If requested row overlaps last_row, auto-shift to last_row + gap_rows + 1 (same column).
    Returns (effective_cell, shifted).
    """
    req_row, req_col = excel_writer._parse_cell(start_cell)
    last = sheet_last_row.get(sheet_name, 0)
    if req_row <= last:
        new_row = last + gap_rows + 1
        return f"{excel_writer._col_letter(req_col)}{new_row}", True
    return start_cell, False


class ExcelReportExecutor:
    def execute(
        self,
        session: Session,
        module: ReportModule,
        template: ReportTemplate,
        mappings: list[ReportSheetMapping],
        execution: ReportExecution,
        params: dict[str, Any] | None = None,
    ) -> ReportExecution:
        _params = params or {}
        work_dir = os.path.join(settings.REPORT_TEMP_DIR, str(execution.id))

        try:
            os.makedirs(work_dir, exist_ok=True)
            execution.status = ExecutionStatusEnum.RUNNING
            execution.started_at = datetime.now(UTC)
            session.add(execution)
            session.commit()

            # Datasources from module
            minio_ds = session.get(DataSource, module.minio_datasource_id)
            sql_ds = session.get(DataSource, module.sql_datasource_id)
            if not minio_ds or not minio_ds.is_active:
                raise ValueError("MinIO datasource not found or inactive")
            if not sql_ds or not sql_ds.is_active:
                raise ValueError("SQL datasource not found or inactive")

            # Resolve buckets: template override → module default
            tpl_bucket = template.template_bucket or module.default_template_bucket
            out_bucket = template.output_bucket or module.default_output_bucket

            # Download template or create blank workbook
            template_local = os.path.join(work_dir, "template.xlsx")
            minio = get_minio_client(minio_ds)
            use_blank = not template.template_path or not template.template_path.strip()

            tpl_size_limit = settings.REPORT_MAX_TEMPLATE_SIZE_MB * 1024 * 1024
            if not use_blank:
                try:
                    download_file(
                        minio, tpl_bucket, template.template_path, template_local,
                        max_size_bytes=tpl_size_limit,
                    )
                except Exception as dl_err:
                    _log.warning("Template download failed, using blank workbook: %s", dl_err)
                    use_blank = True

            if use_blank:
                # Create blank workbook with sheets from mappings
                wb = Workbook()
                default_sheet = wb.active
                sheet_names = {m.sheet_name for m in mappings if m.is_active}
                first = True
                for sn in sorted(sheet_names):
                    if first:
                        default_sheet.title = sn
                        first = False
                    else:
                        wb.create_sheet(sn)
                if first:
                    default_sheet.title = "Sheet1"
                _log.info("Created blank workbook with sheets: %s", list(wb.sheetnames))
            else:
                wb = load_workbook(template_local)
            active_mappings = sorted(
                [m for m in mappings if m.is_active], key=lambda m: m.sort_order
            )
            sql_engine = SQLTemplateEngine()
            chunk_size = settings.REPORT_SQL_CHUNK_SIZE
            max_rows = settings.REPORT_MAX_ROWS_PER_SHEET

            tpl_fmt = template.format_config or {}
            sheet_last_row: dict[str, int] = {}

            total_mappings = len(active_mappings) or 1
            cumulative_rows = 0
            streaming_threshold = settings.REPORT_STREAMING_ROW_THRESHOLD
            warned_streaming = False

            for mapping_idx, mapping in enumerate(active_mappings):
                # Resolve effective start cell (auto-shift on collision)
                eff_cell, shifted = _effective_start(
                    mapping.sheet_name, mapping.start_cell,
                    mapping.gap_rows, sheet_last_row,
                )
                if shifted:
                    _log.info(
                        "Mapping %s: start_cell %s collided with prior writes on sheet %r, "
                        "shifted to %s (gap_rows=%d)",
                        mapping.id, mapping.start_cell, mapping.sheet_name,
                        eff_cell, mapping.gap_rows,
                    )

                # Merge formats: template base ⟵ mapping override
                merged_fmt = _deep_merge(tpl_fmt, mapping.format_config or {}) or {}
                header_fmt = merged_fmt.get("header")
                data_fmt = merged_fmt.get("data")
                col_widths = merged_fmt.get("column_widths")
                do_auto_fit = bool(merged_fmt.get("auto_fit"))
                auto_fit_max = float(merged_fmt.get("auto_fit_max_width") or 50)
                do_wrap_text = bool(merged_fmt.get("wrap_text"))

                start_row, start_col = excel_writer._parse_cell(eff_cell)
                rendered_sql = sql_engine.render(mapping.sql_content, _params)

                if mapping.write_mode == SheetWriteModeEnum.SINGLE:
                    results = execute_sql(sql_ds, rendered_sql, use_pool=True)
                    data = results[0] if results else []
                    if not isinstance(data, int):
                        write_single(
                            wb, mapping.sheet_name, eff_cell, data,
                            data_format=data_fmt,
                        )
                        if data:
                            last = max(sheet_last_row.get(mapping.sheet_name, 0), start_row)
                            sheet_last_row[mapping.sheet_name] = last
                            cumulative_rows += 1
                    # Commit progress after each mapping so async pollers see advancement.
                    pct = min(95, int((mapping_idx + 1) / total_mappings * 95))
                    _update_progress(session, execution, cumulative_rows, pct)
                    continue

                # Chunked pagination for ROWS mode
                sql_upper = rendered_sql.strip().upper()
                user_has_limit = "LIMIT" in sql_upper.split("--")[0].split("/*")[0]

                if user_has_limit:
                    results = execute_sql(sql_ds, rendered_sql, use_pool=True)
                    data = results[0] if results else []
                    if isinstance(data, int):
                        continue
                    if max_rows > 0 and len(data) > max_rows:
                        data = data[:max_rows]
                    rows = write_rows(
                        wb, mapping.sheet_name, eff_cell, data,
                        write_headers=mapping.write_headers,
                        header_format=header_fmt,
                        data_format=data_fmt,
                        column_widths=col_widths,
                        auto_fit=do_auto_fit,
                        auto_fit_max_width=auto_fit_max,
                        wrap_text=do_wrap_text,
                    )
                    if rows:
                        last_row = start_row + rows - 1
                        sheet_last_row[mapping.sheet_name] = max(
                            sheet_last_row.get(mapping.sheet_name, 0), last_row,
                        )
                        cumulative_rows += len(data)
                    pct = min(95, int((mapping_idx + 1) / total_mappings * 95))
                    _update_progress(session, execution, cumulative_rows, pct)
                else:
                    # Auto-paginate
                    total_written = 0
                    headers_written = False
                    offset = 0
                    current_row = start_row
                    # column widths apply once per mapping
                    widths_applied = False

                    while total_written < max_rows:
                        chunk_sql = f"{rendered_sql} LIMIT {chunk_size} OFFSET {offset}"
                        results = execute_sql(sql_ds, chunk_sql, use_pool=True)
                        chunk = results[0] if results else []
                        if isinstance(chunk, int) or not chunk:
                            break

                        if mapping.write_headers and not headers_written:
                            columns = list(chunk[0].keys())
                            ws = wb[mapping.sheet_name]
                            for ci, col_name in enumerate(columns):
                                c = ws.cell(
                                    row=current_row, column=start_col + ci,
                                    value=excel_writer._sanitize_value(col_name),
                                )
                                excel_writer._apply_cell_format(c, header_fmt)
                                if do_wrap_text:
                                    c.alignment = excel_writer.Alignment(
                                        horizontal=c.alignment.horizontal if c.alignment else None,
                                        vertical=c.alignment.vertical if c.alignment else None,
                                        wrap_text=True,
                                    )
                            current_row += 1
                            headers_written = True

                        remaining = max_rows - total_written
                        if len(chunk) > remaining:
                            chunk = chunk[:remaining]

                        rows_written = write_rows(
                            wb, mapping.sheet_name,
                            f"{excel_writer._col_letter(start_col)}{current_row}",
                            chunk, write_headers=False,
                            data_format=data_fmt,
                            column_widths=col_widths if not widths_applied else None,
                            auto_fit=do_auto_fit and not widths_applied,
                            auto_fit_max_width=auto_fit_max,
                            wrap_text=do_wrap_text,
                        )
                        widths_applied = True
                        current_row += rows_written
                        chunk_len = len(chunk)
                        total_written += chunk_len
                        cumulative_rows += chunk_len
                        offset += chunk_size

                        _log.info(
                            "Chunk written: %d rows (total %d) for %s",
                            chunk_len, total_written, mapping.sheet_name,
                        )

                        # Per-chunk progress update (within a mapping, cap at 95%).
                        per_mapping_progress = min(
                            0.95,
                            total_written / max(max_rows, chunk_size * 10),
                        )
                        overall = (mapping_idx + per_mapping_progress) / total_mappings
                        _update_progress(
                            session, execution, cumulative_rows,
                            min(95, int(overall * 95)),
                        )

                        # Memory hygiene on very large runs.
                        if cumulative_rows > streaming_threshold:
                            if not warned_streaming:
                                _log.warning(
                                    "Report exec=%s crossed streaming threshold "
                                    "(%d rows > %d). openpyxl keeps the whole "
                                    "workbook in RAM; consider paginating the "
                                    "template or splitting the report.",
                                    execution.id, cumulative_rows, streaming_threshold,
                                )
                                warned_streaming = True
                            # Drop chunk reference and force collection so the
                            # per-chunk dicts don't linger between iterations.
                            del chunk
                            gc.collect()

                        if chunk_len < chunk_size:
                            break

                    if current_row > start_row:
                        sheet_last_row[mapping.sheet_name] = max(
                            sheet_last_row.get(mapping.sheet_name, 0), current_row - 1,
                        )

            filled_path = os.path.join(work_dir, "filled.xlsx")
            wb.save(filled_path)
            wb.close()

            # Recalc (auto-enable when output_sheet is set)
            output_path = filled_path
            needs_recalc = template.recalc_enabled or bool(template.output_sheet)
            if needs_recalc:
                recalc_timeout = template.recalc_timeout_override or None
                output_path = recalc_workbook(filled_path, timeout=recalc_timeout)

            # Extract sheet (TH2)
            if template.output_sheet:
                extracted_path = os.path.join(work_dir, "output.xlsx")
                extract_sheet(output_path, template.output_sheet, extracted_path)
                output_path = extracted_path

            # Upload
            import re as _re
            timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
            safe_name = _re.sub(r'[^a-z0-9_-]', '_', template.name.lower())
            output_filename = f"{timestamp}_{safe_name}.xlsx"
            object_path = f"{template.output_prefix}{output_filename}"
            out_size_limit = settings.REPORT_MAX_OUTPUT_SIZE_MB * 1024 * 1024
            upload_file(
                minio, out_bucket, object_path, output_path,
                max_size_bytes=out_size_limit,
            )

            execution.status = ExecutionStatusEnum.SUCCESS
            execution.output_minio_path = f"{out_bucket}/{object_path}"
            execution.output_url = None  # Use /download proxy endpoint instead
            execution.processed_rows = cumulative_rows
            execution.progress_pct = 100
            execution.completed_at = datetime.now(UTC)
            session.add(execution)
            session.commit()
            return execution

        except Exception as e:
            _log.error("Report generation failed: %s", e, exc_info=True)
            execution.status = ExecutionStatusEnum.FAILED
            execution.error_message = str(e)[:4000]
            execution.completed_at = datetime.now(UTC)
            session.add(execution)
            session.commit()
            return execution
        finally:
            if os.path.exists(work_dir):
                shutil.rmtree(work_dir, ignore_errors=True)
