"""Write SQL query results into Excel sheets."""
import logging
import re
import uuid
from datetime import date, datetime, time
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

_log = logging.getLogger(__name__)


# Leading chars that Excel treats as the start of a formula. When present in
# string data originating from SQL results they must be neutralised to prevent
# formula/CSV injection attacks (e.g. =HYPERLINK, =cmd|..., @SUM).
_FORMULA_TRIGGERS = ("=", "+", "-", "@", "\t", "\r")


def _escape_formula(val: str) -> str:
    if val and val[0] in _FORMULA_TRIGGERS:
        return "'" + val
    return val


def _sanitize_value(val: object) -> object:
    if val is None:
        return val
    if isinstance(val, str):
        return _escape_formula(val)
    if isinstance(val, (int, float, bool, datetime, date, time)):
        return val
    if isinstance(val, Decimal):
        return float(val)
    if isinstance(val, uuid.UUID):
        return str(val)
    if isinstance(val, bytes):
        return val.hex()
    return _escape_formula(str(val))


def _col_letter(col: int) -> str:
    """Convert 1-indexed column number to letter(s). 1→A, 26→Z, 27→AA."""
    result = ""
    while col > 0:
        col, remainder = divmod(col - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _parse_cell(cell_ref: str) -> tuple[int, int]:
    m = re.match(r"^([A-Za-z]+)(\d+)$", cell_ref.strip())
    if not m:
        raise ValueError(f"Invalid cell reference: {cell_ref!r}")
    col_str = m.group(1).upper()
    row = int(m.group(2))
    col = 0
    for ch in col_str:
        col = col * 26 + (ord(ch) - ord("A") + 1)
    return row, col


# ---------------------------------------------------------------------------
# Format application
# ---------------------------------------------------------------------------


def _build_font(cfg: dict[str, Any] | None, base: Font | None) -> Font | None:
    if not cfg:
        return None
    if base is None:
        base = Font()
    return Font(
        name=cfg.get("name") if cfg.get("name") is not None else base.name,
        size=cfg.get("size") if cfg.get("size") is not None else base.size,
        bold=cfg.get("bold") if cfg.get("bold") is not None else base.bold,
        italic=cfg.get("italic") if cfg.get("italic") is not None else base.italic,
        color=cfg.get("color") if cfg.get("color") is not None else base.color,
    )


def _build_fill(cfg: dict[str, Any] | None) -> PatternFill | None:
    if not cfg:
        return None
    bg = cfg.get("bg_color")
    pattern = cfg.get("pattern") or ("solid" if bg else None)
    if not pattern and not bg:
        return None
    return PatternFill(fill_type=pattern, start_color=bg, end_color=bg)


def _build_border(cfg: dict[str, Any] | None) -> Border | None:
    if not cfg:
        return None
    style = cfg.get("style")
    color = cfg.get("color")
    if not style:
        return None
    side = Side(style=style, color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def _build_alignment(cfg: dict[str, Any] | None) -> Alignment | None:
    if not cfg:
        return None
    kwargs: dict[str, Any] = {}
    if (v := cfg.get("horizontal")) is not None:
        kwargs["horizontal"] = v
    if (v := cfg.get("vertical")) is not None:
        kwargs["vertical"] = v
    if (v := cfg.get("wrap_text")) is not None:
        kwargs["wrap_text"] = v
    if not kwargs:
        return None
    return Alignment(**kwargs)


def _apply_cell_format(cell: Any, fmt: dict[str, Any] | None) -> None:
    """Apply a CellFormat dict to an openpyxl cell (partial/merge with existing)."""
    if not fmt:
        return
    font = _build_font(fmt.get("font"), cell.font)
    if font is not None:
        cell.font = font
    fill = _build_fill(fmt.get("fill"))
    if fill is not None:
        cell.fill = fill
    border = _build_border(fmt.get("border"))
    if border is not None:
        cell.border = border
    alignment = _build_alignment(fmt.get("alignment"))
    if alignment is not None:
        cell.alignment = alignment
    nf = fmt.get("number_format")
    if nf:
        cell.number_format = nf


def _apply_column_widths(ws: Any, widths: dict[str, float] | None) -> None:
    if not widths:
        return
    for col, w in widths.items():
        try:
            ws.column_dimensions[col.upper()].width = float(w)
        except Exception as e:
            _log.warning("Invalid column width %s=%s: %s", col, w, e)


# ---------------------------------------------------------------------------
# Write ops
# ---------------------------------------------------------------------------


def _auto_fit_columns(
    ws: Any, start_col: int, columns: list[str],
    data: list[dict], has_headers: bool, max_width: float,
) -> None:
    """Set column widths based on content length. Considers header + data values."""
    for ci, col_name in enumerate(columns):
        best = len(str(col_name)) if has_headers else 0
        for row_data in data:
            val = row_data.get(col_name)
            if val is not None:
                length = max(len(line) for line in str(val).split("\n"))
                best = max(best, length)
        width = min(best + 2, max_width)
        letter = _col_letter(start_col + ci)
        ws.column_dimensions[letter].width = max(width, 8)


def write_rows(
    wb: Workbook, sheet_name: str, start_cell: str,
    data: list[dict], *,
    write_headers: bool = False,
    header_format: dict[str, Any] | None = None,
    data_format: dict[str, Any] | None = None,
    column_widths: dict[str, float] | None = None,
    auto_fit: bool = False,
    auto_fit_max_width: float = 50,
    wrap_text: bool = False,
) -> int:
    """Write rows starting at start_cell. Returns number of rows written (incl header)."""
    if not data:
        return 0
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not found")
    ws = wb[sheet_name]
    start_row, start_col = _parse_cell(start_cell)
    columns = list(data[0].keys())
    current_row = start_row

    _apply_column_widths(ws, column_widths)

    if write_headers:
        for ci, col_name in enumerate(columns):
            c = ws.cell(row=current_row, column=start_col + ci, value=_sanitize_value(col_name))
            _apply_cell_format(c, header_format)
            if wrap_text:
                c.alignment = Alignment(
                    horizontal=c.alignment.horizontal if c.alignment else None,
                    vertical=c.alignment.vertical if c.alignment else None,
                    wrap_text=True,
                )
        current_row += 1
    for row_data in data:
        for ci, col_name in enumerate(columns):
            c = ws.cell(row=current_row, column=start_col + ci, value=_sanitize_value(row_data.get(col_name)))
            _apply_cell_format(c, data_format)
            if wrap_text:
                c.alignment = Alignment(
                    horizontal=c.alignment.horizontal if c.alignment else None,
                    vertical=c.alignment.vertical if c.alignment else None,
                    wrap_text=True,
                )
        current_row += 1

    if auto_fit and not column_widths:
        _auto_fit_columns(ws, start_col, columns, data, write_headers, auto_fit_max_width)

    return current_row - start_row


def write_single(
    wb: Workbook, sheet_name: str, start_cell: str, data: list[dict],
    *, data_format: dict[str, Any] | None = None,
) -> None:
    if not data:
        return
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not found")
    ws = wb[sheet_name]
    row, col = _parse_cell(start_cell)
    first_key = list(data[0].keys())[0]
    c = ws.cell(row=row, column=col, value=_sanitize_value(data[0][first_key]))
    _apply_cell_format(c, data_format)
