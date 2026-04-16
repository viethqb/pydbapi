"""Extract sheets from a workbook with values + styles (theme preserved)."""
import logging

from openpyxl import load_workbook

_log = logging.getLogger(__name__)


def extract_sheet(
    source_path: str,
    sheet_names: str | list[str],
    output_path: str,
) -> None:
    """Extract one or more sheets, deleting the rest.

    *sheet_names* can be a single name, a list, or a comma-separated string.
    """
    if isinstance(sheet_names, str):
        keep = {s.strip() for s in sheet_names.split(",") if s.strip()}
    else:
        keep = set(sheet_names)

    wb = load_workbook(source_path, data_only=True)

    missing = keep - set(wb.sheetnames)
    if missing:
        raise ValueError(f"Sheet(s) not found: {', '.join(sorted(missing))}")

    for name in wb.sheetnames:
        if name not in keep:
            del wb[name]

    wb.save(output_path)
    wb.close()
